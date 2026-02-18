"""Excel export builder for statement detail pages.

This module owns XLSX rendering concerns so the Flask route file can stay focused on
request flow and context assembly.
"""

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from core.statement_row_palette import STATEMENT_ROW_PALETTE
from utils.statement_rows import format_item_type_label, xero_ids_for_row


def _build_excel_headers(display_headers: list[str]) -> tuple[list[tuple[str, str]], list[str]]:
    """Build label pairs and the Excel header row.

    Args:
        display_headers: Ordered display headers from statement config.

    Returns:
        Tuple of (header_labels, excel_headers).
    """
    header_labels: list[tuple[str, str]] = []
    statement_headers: list[str] = []
    xero_headers: list[str] = []

    for header in display_headers:
        label = (header or "").replace("_", " ").strip()
        label = label[0].upper() + label[1:] if label else header or ""
        header_labels.append((header, label))
        statement_headers.append(f"Statement {label}")
        xero_headers.append(f"Xero {label}")

    excel_headers = ["Type", *statement_headers, *xero_headers, "Xero Link", "Status"]
    return header_labels, excel_headers


def _excel_fill_from_hex(color_hex: str) -> PatternFill:
    """Create an Excel solid fill from a #RRGGBB color.

    Args:
        color_hex: Hex color in #RRGGBB format.

    Returns:
        OpenPyXL fill object for that color.
    """
    return PatternFill(fill_type="solid", fgColor=color_hex.lstrip("#"))


def _build_excel_state_fills() -> dict[str, dict[str, PatternFill]]:
    """Build normal/completed row fills for each statement state.

    Args:
        None.

    Returns:
        Nested fill map keyed by state then variant.
    """
    fills: dict[str, dict[str, PatternFill]] = {}
    for state, variants in STATEMENT_ROW_PALETTE.items():
        fills[state] = {"normal": _excel_fill_from_hex(variants["normal"]["background"]), "completed": _excel_fill_from_hex(variants["completed"]["background"])}
    return fills


def _add_excel_legend(workbook: Workbook, *, state_fills: dict[str, dict[str, PatternFill]], mismatch_border: Border) -> None:
    """Add a legend sheet describing statement row styles.

    Args:
        workbook: Workbook being exported.
        state_fills: Nested fills keyed by state and variant.
        mismatch_border: Border style for matched-row cell mismatches.

    Returns:
        None.
    """
    legend = workbook.create_sheet(title="Legend")
    legend.column_dimensions["A"].width = 35
    legend.column_dimensions["B"].width = 18
    legend.append(["Legend", ""])
    legend["A1"].font = Font(bold=True)

    legend_rows = [
        ("Match", "match", "normal"),
        ("Match (Completed)", "match", "completed"),
        ("Mismatch", "mismatch", "normal"),
        ("Mismatch (Completed)", "mismatch", "completed"),
        ("Flagged anomaly", "anomaly", "normal"),
        ("Flagged anomaly (Completed)", "anomaly", "completed"),
    ]

    for label, state, variant in legend_rows:
        legend.append([label, ""])
        legend[f"B{legend.max_row}"].fill = state_fills[state][variant]

    legend.append(["Cell mismatch (matched rows)", ""])
    legend[f"B{legend.max_row}"].border = mismatch_border


def _status_for_excel_row(item: Any, item_status_map: dict[str, bool]) -> tuple[str, bool]:
    """Return the status label and completion flag for an item.

    Args:
        item: Statement item payload.
        item_status_map: Completion map keyed by statement item id.

    Returns:
        Tuple of (status_label, is_item_completed).
    """
    statement_item_id = item.get("statement_item_id") if isinstance(item, dict) else None
    if statement_item_id:
        is_item_completed = item_status_map.get(statement_item_id, False)
        status_label = "Completed" if is_item_completed else "Incomplete"
        return status_label, is_item_completed
    return "", False


def _build_excel_row_values(header_labels: list[tuple[str, str]], left_row: dict[str, Any], right_row: dict[str, Any], item_types: list[str], idx: int) -> list[Any]:
    """Build Excel row values from statement/xero data.

    Args:
        header_labels: Source headers and display labels.
        left_row: Statement-side row values.
        right_row: Xero-side row values.
        item_types: Row-level item type values.
        idx: Current row index.

    Returns:
        Row values for worksheet append().
    """
    item_type = item_types[idx] if idx < len(item_types) else ""
    row_values: list[Any] = [format_item_type_label(item_type)]
    for src_header, _ in header_labels:
        left_value = left_row.get(src_header, "") if isinstance(left_row, dict) else ""
        row_values.append("" if left_value is None else left_value)

    for src_header, _ in header_labels:
        right_value = right_row.get(src_header, "") if isinstance(right_row, dict) else ""
        row_values.append("" if right_value is None else right_value)

    return row_values


def _is_anomalous_item(item: Any) -> bool:
    """Return True when the item has anomaly flags.

    Args:
        item: Statement item payload.

    Returns:
        True when known anomaly flags are present.
    """
    raw_flags = item.get("_flags") if isinstance(item, dict) else None
    flag_list = raw_flags if isinstance(raw_flags, list) else []
    return any(isinstance(flag, str) and flag.strip() in {"ml-outlier", "invalid-date"} for flag in flag_list)


def _row_state_for_item(item: Any, row_match: bool) -> str:
    """Return the semantic statement row state.

    Args:
        item: Statement item payload.
        row_match: Whether the row values match.

    Returns:
        Row state key used by shared UI/Excel color mapping.
    """
    if _is_anomalous_item(item):
        return "anomaly"
    return "match" if row_match else "mismatch"


def _apply_row_fill(worksheet: Any, *, current_row: int, total_columns: int, fill: PatternFill) -> None:
    """Apply row coloring to a worksheet row.

    Args:
        worksheet: Worksheet being exported.
        current_row: Target row number.
        total_columns: Number of visible columns to fill.
        fill: Fill style for the row.

    Returns:
        None.
    """
    for col in range(1, total_columns + 1):
        cell = worksheet.cell(row=current_row, column=col)
        cell.fill = fill


def _apply_divider_borders(worksheet: Any, *, current_row: int, statement_end_col: int, xero_start_col: int, divider_side: Side) -> None:
    """Apply divider borders between statement and Xero sections.

    Args:
        worksheet: Worksheet being exported.
        current_row: Target row number.
        statement_end_col: Last statement column index.
        xero_start_col: First Xero column index.
        divider_side: Border side style.

    Returns:
        None.
    """
    worksheet.cell(row=current_row, column=statement_end_col).border = Border(right=divider_side)
    worksheet.cell(row=current_row, column=xero_start_col).border = Border(left=divider_side)


def _apply_mismatch_borders(
    worksheet: Any,
    *,
    header_labels: list[tuple[str, str]],
    comparisons: list[Any],
    current_row: int,
    statement_end_col: int,
    xero_start_col: int,
    mismatch_border: Border,
    mismatch_side: Side,
    divider_side: Side,
) -> None:
    """Apply per-cell mismatch borders for matched rows.

    Args:
        worksheet: Worksheet being exported.
        header_labels: Source headers and display labels.
        comparisons: Per-cell comparison values for the row.
        current_row: Target row number.
        statement_end_col: Last statement column index.
        xero_start_col: First Xero column index.
        mismatch_border: Default mismatch border style.
        mismatch_side: Side style for mismatch edges.
        divider_side: Side style for statement/Xero split borders.

    Returns:
        None.
    """
    col_count = len(header_labels)
    for col_idx, comparison in enumerate(comparisons[:col_count]):
        if getattr(comparison, "matches", True):
            continue
        for target_col in (2 + col_idx, 2 + col_count + col_idx):
            cell = worksheet.cell(row=current_row, column=target_col)
            if target_col == statement_end_col:
                cell.border = Border(left=mismatch_side, right=divider_side, top=mismatch_side, bottom=mismatch_side)
            elif target_col == xero_start_col:
                cell.border = Border(left=divider_side, right=mismatch_side, top=mismatch_side, bottom=mismatch_side)
            else:
                cell.border = mismatch_border


def _parse_date_value(value: Any) -> date | None:
    """Parse a date value from a record field.

    Args:
        value: Raw date value.

    Returns:
        Parsed date or None.
    """
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _format_date_segment(earliest_date: date | None, latest_date: date | None) -> str:
    """Return the filename date segment from the parsed dates.

    Args:
        earliest_date: Earliest statement item date.
        latest_date: Latest statement item date.

    Returns:
        Filename-friendly date segment.
    """
    if earliest_date and latest_date:
        if earliest_date == latest_date:
            return earliest_date.strftime("%Y-%m-%d")
        return f"{earliest_date.strftime('%Y-%m-%d')}_{latest_date.strftime('%Y-%m-%d')}"
    if latest_date or earliest_date:
        chosen = latest_date or earliest_date
        return chosen.strftime("%Y-%m-%d") if chosen else ""
    return ""


def _contact_segment(record: dict[str, Any], statement_id: str) -> str:
    """Return the filename contact segment.

    Args:
        record: Statement record payload.
        statement_id: Statement id.

    Returns:
        Safe contact-name segment for file naming.
    """
    contact_name = record.get("ContactName") if isinstance(record, dict) else ""
    return secure_filename(str(contact_name or "").strip()) or f"statement_{statement_id}"


def _append_excel_rows(
    worksheet: Any,
    *,
    header_labels: list[tuple[str, str]],
    excel_headers: list[str],
    rows_by_header: list[dict[str, Any]],
    right_rows_by_header: list[dict[str, Any]],
    row_comparisons: list[list[Any]],
    row_matches: list[bool],
    item_types: list[str],
    items: list[Any],
    item_number_header: str | None,
    matched_invoice_to_statement_item: dict[str, Any],
    item_status_map: dict[str, bool],
    statement_col_count: int,
    statement_end_col: int,
    xero_start_col: int,
    state_fills: dict[str, dict[str, PatternFill]],
    mismatch_border: Border,
    mismatch_side: Side,
    divider_side: Side,
) -> int:
    """Append rows to the Excel worksheet and return row count.

    This includes a hyperlink cell for the Xero Link column when available.

    Args:
        worksheet: Worksheet being exported.
        header_labels: Source headers and display labels.
        excel_headers: Visible worksheet headers.
        rows_by_header: Statement rows keyed by header.
        right_rows_by_header: Xero rows keyed by header.
        row_comparisons: Per-cell comparison data by row.
        row_matches: Per-row match flags.
        item_types: Row-level statement item types.
        items: Raw statement items payload.
        item_number_header: Statement header used for Xero match lookup.
        matched_invoice_to_statement_item: Statement->Xero match map.
        item_status_map: Completion map keyed by statement item id.
        statement_col_count: Number of statement-side columns.
        statement_end_col: Last statement column index.
        xero_start_col: First Xero column index.
        state_fills: Nested fill map keyed by state and variant.
        mismatch_border: Border style for mismatched cells.
        mismatch_side: Border side used on mismatch edges.
        divider_side: Border side used on statement/Xero split columns.

    Returns:
        Number of data rows appended.
    """
    row_count = max(len(rows_by_header), len(right_rows_by_header))
    try:
        link_col = excel_headers.index("Xero Link") + 1
    except ValueError:
        link_col = None

    for idx in range(row_count):
        left_row = rows_by_header[idx] if idx < len(rows_by_header) else {}
        right_row = right_rows_by_header[idx] if idx < len(right_rows_by_header) else {}
        item = items[idx] if idx < len(items) else {}

        status_label, is_item_completed = _status_for_excel_row(item, item_status_map)
        row_values = _build_excel_row_values(header_labels, left_row, right_row, item_types, idx)
        xero_invoice_id, xero_credit_note_id = xero_ids_for_row(item_number_header, left_row, matched_invoice_to_statement_item)
        if xero_credit_note_id:
            xero_link = f"https://go.xero.com/AccountsPayable/ViewCreditNote.aspx?creditNoteID={xero_credit_note_id}"
        elif xero_invoice_id:
            xero_link = f"https://go.xero.com/AccountsPayable/View.aspx?InvoiceID={xero_invoice_id}"
        else:
            xero_link = ""

        # Providing status in the sheet lets users filter finished work out quickly.
        row_values.append("Link" if xero_link else "")
        row_values.append(status_label)
        worksheet.append(row_values)
        current_row = worksheet.max_row
        if xero_link and link_col:
            link_cell = worksheet.cell(row=current_row, column=link_col)
            link_cell.hyperlink = xero_link

        row_match = row_matches[idx] if idx < len(row_matches) else False
        row_state = _row_state_for_item(item, row_match)
        fill_variant = "completed" if is_item_completed else "normal"
        fill = state_fills[row_state][fill_variant]
        _apply_row_fill(worksheet, current_row=current_row, total_columns=len(excel_headers), fill=fill)

        if statement_col_count:
            _apply_divider_borders(worksheet, current_row=current_row, statement_end_col=statement_end_col, xero_start_col=xero_start_col, divider_side=divider_side)

        if row_match and idx < len(row_comparisons):
            comparisons = row_comparisons[idx] or []
            _apply_mismatch_borders(
                worksheet,
                header_labels=header_labels,
                comparisons=comparisons,
                current_row=current_row,
                statement_end_col=statement_end_col,
                xero_start_col=xero_start_col,
                mismatch_border=mismatch_border,
                mismatch_side=mismatch_side,
                divider_side=divider_side,
            )
    return row_count


def build_statement_excel_payload(
    *,
    display_headers: list[str],
    rows_by_header: list[dict[str, Any]],
    right_rows_by_header: list[dict[str, Any]],
    row_comparisons: list[list[Any]],
    row_matches: list[bool],
    item_types: list[str],
    items: list[Any],
    item_number_header: str | None,
    matched_invoice_to_statement_item: dict[str, Any],
    item_status_map: dict[str, bool],
    record: dict[str, Any],
    statement_id: str,
) -> tuple[bytes, str, int]:
    """Build XLSX bytes and filename metadata for statement export.

    Args:
        display_headers: Statement display headers.
        rows_by_header: Statement rows keyed by header.
        right_rows_by_header: Xero rows keyed by header.
        row_comparisons: Per-cell comparison results.
        row_matches: Per-row match flags.
        item_types: Item type labels per row.
        items: Statement items payload.
        item_number_header: Header used to map Xero links.
        matched_invoice_to_statement_item: Matched Xero invoice map.
        item_status_map: Statement item completion flags.
        record: Statement metadata record.
        statement_id: Statement identifier.

    Returns:
        Tuple of (excel_payload_bytes, download_filename, row_count).
    """
    header_labels, excel_headers = _build_excel_headers(display_headers)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Statement"
    worksheet.append(excel_headers)

    state_fills = _build_excel_state_fills()
    mismatch_side = Side(style="thin", color="D8A0A0")
    mismatch_border = Border(left=mismatch_side, right=mismatch_side, top=mismatch_side, bottom=mismatch_side)
    divider_side = Side(style="medium", color="808080")
    statement_col_count = len(header_labels)
    statement_end_col = 1 + statement_col_count
    xero_start_col = statement_end_col + 1

    _add_excel_legend(workbook, state_fills=state_fills, mismatch_border=mismatch_border)

    if statement_col_count:
        worksheet.cell(row=1, column=statement_end_col).border = Border(right=divider_side)
        worksheet.cell(row=1, column=xero_start_col).border = Border(left=divider_side)

    # Pylint's duplicate-code check compares this pass-through block with app.py.
    # Keeping the call explicit avoids hidden argument coupling during future changes.
    # pylint: disable=duplicate-code
    row_count = _append_excel_rows(
        worksheet,
        header_labels=header_labels,
        excel_headers=excel_headers,
        rows_by_header=rows_by_header,
        right_rows_by_header=right_rows_by_header,
        row_comparisons=row_comparisons,
        row_matches=row_matches,
        item_types=item_types,
        items=items,
        item_number_header=item_number_header,
        matched_invoice_to_statement_item=matched_invoice_to_statement_item,
        item_status_map=item_status_map,
        statement_col_count=statement_col_count,
        statement_end_col=statement_end_col,
        xero_start_col=xero_start_col,
        state_fills=state_fills,
        mismatch_border=mismatch_border,
        mismatch_side=mismatch_side,
        divider_side=divider_side,
    )
    # pylint: enable=duplicate-code

    header_font = Font(bold=True)
    for col_idx in range(1, len(excel_headers) + 1):
        worksheet.cell(row=1, column=col_idx).font = header_font

    worksheet.freeze_panes = "A2"
    last_row = max(row_count + 1, 1)
    last_column = get_column_letter(len(excel_headers))
    worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    width_overrides = {"Type": 8, "Status": 12, "Xero Link": 12}
    for col_idx, header in enumerate(excel_headers, start=1):
        width = width_overrides.get(header)
        if width is None:
            width = min(max(len(header) + 2, 14), 30)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    excel_payload = output.getvalue()
    output.close()

    earliest_date = _parse_date_value(record.get("EarliestItemDate"))
    latest_date = _parse_date_value(record.get("LatestItemDate"))
    date_segment = _format_date_segment(earliest_date, latest_date)
    contact_segment = _contact_segment(record, statement_id)
    parts = [contact_segment]
    if date_segment:
        parts.append(date_segment)
    download_name = "_".join(parts) + "_export.xlsx"
    return excel_payload, download_name, row_count
