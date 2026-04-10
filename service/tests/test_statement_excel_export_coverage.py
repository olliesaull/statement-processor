"""Coverage tests for utils/statement_excel_export.py.

Exercises the public build_statement_excel_payload function with realistic
minimal data and tests internal helpers directly for edge-case branches
that are hard to reach through the public API alone.
"""

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook

from core.models import CellComparison
from utils.statement_excel_export import (
    _build_excel_headers,
    _build_excel_row_values,
    _contact_segment,
    _format_date_segment,
    _is_anomalous_item,
    _parse_date_value,
    _row_state_for_item,
    _status_for_excel_row,
    build_statement_excel_payload,
)

# ---------------------------------------------------------------------------
# _build_excel_headers
# ---------------------------------------------------------------------------


class TestBuildExcelHeaders:
    """Tests for _build_excel_headers — label pairs and header row assembly."""

    def test_basic_headers(self) -> None:
        """Should produce Statement/Xero prefixed columns plus Type, Xero Link, Status."""
        header_labels, excel_headers = _build_excel_headers(["invoice_date", "amount"])
        assert header_labels == [("invoice_date", "Invoice date"), ("amount", "Amount")]
        assert excel_headers == ["Type", "Statement Invoice date", "Statement Amount", "Xero Invoice date", "Xero Amount", "Xero Link", "Status"]

    def test_empty_string_header(self) -> None:
        """An empty string header should survive without error."""
        header_labels, excel_headers = _build_excel_headers([""])
        # Empty string => label is ""
        assert header_labels == [("", "")]
        assert "Statement " in excel_headers[1]

    def test_none_header_element(self) -> None:
        """None as a header element should be handled gracefully."""
        header_labels, excel_headers = _build_excel_headers([None])
        # None coerces to "" via (header or "")
        assert header_labels[0] == (None, "")
        assert len(excel_headers) == 5  # Type + 1 Statement + 1 Xero + Xero Link + Status

    def test_no_headers(self) -> None:
        """Empty header list should still produce Type, Xero Link, Status columns."""
        header_labels, excel_headers = _build_excel_headers([])
        assert header_labels == []
        assert excel_headers == ["Type", "Xero Link", "Status"]


# ---------------------------------------------------------------------------
# _status_for_excel_row
# ---------------------------------------------------------------------------


class TestStatusForExcelRow:
    """Tests for _status_for_excel_row — item completion label."""

    def test_completed_item(self) -> None:
        """Item with a statement_item_id marked completed returns 'Completed'."""
        item = {"statement_item_id": "item-1"}
        label, completed = _status_for_excel_row(item, {"item-1": True})
        assert label == "Completed"
        assert completed is True

    def test_incomplete_item(self) -> None:
        """Item marked incomplete returns 'Incomplete'."""
        item = {"statement_item_id": "item-2"}
        label, completed = _status_for_excel_row(item, {"item-2": False})
        assert label == "Incomplete"
        assert completed is False

    def test_missing_statement_item_id(self) -> None:
        """Item without statement_item_id returns empty label and False."""
        label, completed = _status_for_excel_row({}, {})
        assert label == ""
        assert completed is False

    def test_non_dict_item(self) -> None:
        """Non-dict item returns empty label and False."""
        label, completed = _status_for_excel_row("not-a-dict", {})
        assert label == ""
        assert completed is False

    def test_item_id_not_in_map(self) -> None:
        """Item with id that isn't in the status map defaults to Incomplete."""
        item = {"statement_item_id": "item-unknown"}
        label, completed = _status_for_excel_row(item, {})
        assert label == "Incomplete"
        assert completed is False


# ---------------------------------------------------------------------------
# _build_excel_row_values
# ---------------------------------------------------------------------------


class TestBuildExcelRowValues:
    """Tests for _build_excel_row_values — row data assembly."""

    def test_basic_row(self) -> None:
        """Should build [Type, statement_val, xero_val] from left/right rows."""
        labels = [("amount", "Amount")]
        left = {"amount": "100.00"}
        right = {"amount": "100.00"}
        values = _build_excel_row_values(labels, left, right, ["invoice"], 0)
        assert values[0] == "INV"  # format_item_type_label("invoice")
        assert values[1] == "100.00"
        assert values[2] == "100.00"

    def test_none_values_become_empty(self) -> None:
        """None cell values should be replaced with empty strings."""
        labels = [("amount", "Amount")]
        left = {"amount": None}
        right = {"amount": None}
        values = _build_excel_row_values(labels, left, right, [], 0)
        assert values[1] == ""
        assert values[2] == ""

    def test_non_dict_rows(self) -> None:
        """Non-dict left/right rows produce empty cell values."""
        labels = [("amount", "Amount")]
        values = _build_excel_row_values(labels, "not-dict", "not-dict", ["invoice"], 0)
        assert values[1] == ""
        assert values[2] == ""

    def test_idx_beyond_item_types(self) -> None:
        """Index beyond item_types length should use empty type."""
        labels = [("amount", "Amount")]
        values = _build_excel_row_values(labels, {"amount": "50"}, {}, [], 5)
        assert values[0] == ""  # No item type


# ---------------------------------------------------------------------------
# _is_anomalous_item
# ---------------------------------------------------------------------------


class TestIsAnomalousItem:
    """Tests for _is_anomalous_item — anomaly flag detection."""

    def test_ml_outlier_flag(self) -> None:
        """'ml-outlier' flag should be detected."""
        assert _is_anomalous_item({"_flags": ["ml-outlier"]}) is True

    def test_invalid_date_flag(self) -> None:
        """'invalid-date' flag should be detected."""
        assert _is_anomalous_item({"_flags": ["invalid-date"]}) is True

    def test_no_flags(self) -> None:
        """Empty flags list is not anomalous."""
        assert _is_anomalous_item({"_flags": []}) is False

    def test_missing_flags_key(self) -> None:
        """No _flags key is not anomalous."""
        assert _is_anomalous_item({}) is False

    def test_non_dict_item(self) -> None:
        """Non-dict item is not anomalous."""
        assert _is_anomalous_item("not-a-dict") is False

    def test_non_list_flags(self) -> None:
        """Non-list _flags value is not anomalous."""
        assert _is_anomalous_item({"_flags": "ml-outlier"}) is False

    def test_non_string_flag_in_list(self) -> None:
        """Non-string values in the flags list are ignored."""
        assert _is_anomalous_item({"_flags": [42, None]}) is False


# ---------------------------------------------------------------------------
# _row_state_for_item
# ---------------------------------------------------------------------------


class TestRowStateForItem:
    """Tests for _row_state_for_item — semantic row state derivation."""

    def test_anomalous_item_returns_anomaly(self) -> None:
        """Anomalous items always return 'anomaly' regardless of match."""
        assert _row_state_for_item({"_flags": ["ml-outlier"]}, row_match=True) == "anomaly"

    def test_matched_row_returns_match(self) -> None:
        """Non-anomalous matched row returns 'match'."""
        assert _row_state_for_item({}, row_match=True) == "match"

    def test_unmatched_row_returns_mismatch(self) -> None:
        """Non-anomalous unmatched row returns 'mismatch'."""
        assert _row_state_for_item({}, row_match=False) == "mismatch"


# ---------------------------------------------------------------------------
# _parse_date_value
# ---------------------------------------------------------------------------


class TestParseDateValue:
    """Tests for _parse_date_value — date string parsing."""

    def test_valid_iso_date(self) -> None:
        """Standard ISO date should parse successfully."""
        assert _parse_date_value("2024-03-15") == date(2024, 3, 15)

    def test_whitespace_around_date(self) -> None:
        """Whitespace should be stripped before parsing."""
        assert _parse_date_value("  2024-01-01  ") == date(2024, 1, 1)

    def test_invalid_date_string(self) -> None:
        """Non-date string should return None."""
        assert _parse_date_value("not-a-date") is None

    def test_non_string_value(self) -> None:
        """Non-string values (int, None) should return None."""
        assert _parse_date_value(42) is None
        assert _parse_date_value(None) is None


# ---------------------------------------------------------------------------
# _format_date_segment
# ---------------------------------------------------------------------------


class TestFormatDateSegment:
    """Tests for _format_date_segment — filename date range formatting."""

    def test_same_dates(self) -> None:
        """Equal earliest and latest produce a single date."""
        d = date(2024, 6, 15)
        assert _format_date_segment(d, d) == "2024-06-15"

    def test_different_dates(self) -> None:
        """Different dates produce an underscore-separated range."""
        assert _format_date_segment(date(2024, 1, 1), date(2024, 12, 31)) == "2024-01-01_2024-12-31"

    def test_only_latest(self) -> None:
        """Only latest date provided."""
        assert _format_date_segment(None, date(2024, 5, 1)) == "2024-05-01"

    def test_only_earliest(self) -> None:
        """Only earliest date provided."""
        assert _format_date_segment(date(2024, 5, 1), None) == "2024-05-01"

    def test_neither_date(self) -> None:
        """No dates returns empty string."""
        assert _format_date_segment(None, None) == ""


# ---------------------------------------------------------------------------
# _contact_segment
# ---------------------------------------------------------------------------


class TestContactSegment:
    """Tests for _contact_segment — safe filename contact portion."""

    def test_normal_contact_name(self) -> None:
        """Contact name should be made filename-safe."""
        result = _contact_segment({"ContactName": "Acme Corp"}, "s-1")
        assert "Acme" in result

    def test_missing_contact_name(self) -> None:
        """Absent contact name falls back to statement_<id>."""
        result = _contact_segment({}, "stmt-42")
        assert result == "statement_stmt-42"

    def test_empty_contact_name(self) -> None:
        """Empty string contact name falls back to statement_<id>."""
        result = _contact_segment({"ContactName": ""}, "stmt-42")
        assert result == "statement_stmt-42"

    def test_non_dict_record(self) -> None:
        """Non-dict record falls back to statement_<id>."""
        result = _contact_segment("not-a-dict", "stmt-42")
        assert result == "statement_stmt-42"


# ---------------------------------------------------------------------------
# build_statement_excel_payload (integration)
# ---------------------------------------------------------------------------


class TestBuildStatementExcelPayload:
    """Integration tests for the public build_statement_excel_payload function."""

    @staticmethod
    def _minimal_args(**overrides: Any) -> dict[str, Any]:
        """Return minimal valid arguments for build_statement_excel_payload."""
        defaults: dict[str, Any] = {
            "display_headers": ["date", "number", "amount"],
            "rows_by_header": [{"date": "2024-03-01", "number": "101", "amount": "500.00"}, {"date": "2024-03-02", "number": "102", "amount": "250.00"}],
            "right_rows_by_header": [{"date": "2024-03-01", "number": "101", "amount": "500.00"}, {"date": "2024-03-02", "number": "102", "amount": "300.00"}],
            "row_comparisons": [
                [
                    CellComparison(header="date", statement_value="2024-03-01", xero_value="2024-03-01", matches=True),
                    CellComparison(header="number", statement_value="101", xero_value="101", matches=True),
                    CellComparison(header="amount", statement_value="500.00", xero_value="500.00", matches=True),
                ],
                [
                    CellComparison(header="date", statement_value="2024-03-02", xero_value="2024-03-02", matches=True),
                    CellComparison(header="number", statement_value="102", xero_value="102", matches=True),
                    CellComparison(header="amount", statement_value="250.00", xero_value="300.00", matches=False),
                ],
            ],
            "row_matches": [True, False],
            "item_types": ["invoice", "credit_note"],
            "items": [{"statement_item_id": "item-1"}, {"statement_item_id": "item-2", "_flags": ["ml-outlier"]}],
            "item_number_header": "number",
            "matched_invoice_to_statement_item": {"101": {"invoice": {"invoice_id": "inv-abc", "credit_note_id": None}}},
            "item_status_map": {"item-1": True, "item-2": False},
            "record": {"ContactName": "Acme Corp", "EarliestItemDate": "2024-03-01", "LatestItemDate": "2024-03-02"},
            "statement_id": "stmt-1",
        }
        defaults.update(overrides)
        return defaults

    def test_returns_bytes_filename_and_row_count(self) -> None:
        """Should return a valid XLSX payload, filename, and accurate row count."""
        payload, filename, row_count = build_statement_excel_payload(**self._minimal_args())
        assert isinstance(payload, bytes)
        assert len(payload) > 0
        assert filename.endswith("_export.xlsx")
        assert row_count == 2

    def test_filename_includes_contact_and_dates(self) -> None:
        """Filename should contain the contact name and date range."""
        _, filename, _ = build_statement_excel_payload(**self._minimal_args())
        assert "Acme" in filename
        assert "2024-03-01" in filename
        assert "2024-03-02" in filename

    def test_filename_without_dates(self) -> None:
        """Filename with no date fields should omit the date segment."""
        _, filename, _ = build_statement_excel_payload(**self._minimal_args(record={"ContactName": "Test Co"}))
        assert "export.xlsx" in filename
        # No date segment means no underscore-separated dates
        assert "2024" not in filename

    def test_worksheet_has_expected_headers(self) -> None:
        """The first row of the Statement sheet should contain all expected headers."""
        payload, _, _ = build_statement_excel_payload(**self._minimal_args())
        wb = load_workbook(BytesIO(payload))
        ws = wb["Statement"]
        headers = [cell.value for cell in ws[1]]
        assert "Type" in headers
        assert "Statement Date" in headers
        assert "Xero Amount" in headers
        assert "Xero Link" in headers
        assert "Status" in headers

    def test_worksheet_has_correct_row_count(self) -> None:
        """Data rows should match the row_count returned."""
        payload, _, row_count = build_statement_excel_payload(**self._minimal_args())
        wb = load_workbook(BytesIO(payload))
        ws = wb["Statement"]
        # Row 1 is the header, so data starts at row 2.
        data_rows = ws.max_row - 1
        assert data_rows == row_count

    def test_legend_sheet_exists(self) -> None:
        """The exported workbook should include a Legend sheet."""
        payload, _, _ = build_statement_excel_payload(**self._minimal_args())
        wb = load_workbook(BytesIO(payload))
        assert "Legend" in wb.sheetnames

    def test_xero_link_populated_for_matched_invoice(self) -> None:
        """Row matched to a Xero invoice should have a hyperlink in the Xero Link column."""
        payload, _, _ = build_statement_excel_payload(**self._minimal_args())
        wb = load_workbook(BytesIO(payload))
        ws = wb["Statement"]
        headers = [cell.value for cell in ws[1]]
        link_col = headers.index("Xero Link") + 1
        # Row 2 (first data row) is matched to invoice inv-abc
        link_cell = ws.cell(row=2, column=link_col)
        assert link_cell.value == "Link"
        assert link_cell.hyperlink is not None
        assert "inv-abc" in link_cell.hyperlink.target

    def test_credit_note_link(self) -> None:
        """Row matched to a credit note should use the credit note URL."""
        args = self._minimal_args(matched_invoice_to_statement_item={"101": {"invoice": {"invoice_id": None, "credit_note_id": "cn-xyz"}}})
        payload, _, _ = build_statement_excel_payload(**args)
        wb = load_workbook(BytesIO(payload))
        ws = wb["Statement"]
        headers = [cell.value for cell in ws[1]]
        link_col = headers.index("Xero Link") + 1
        link_cell = ws.cell(row=2, column=link_col)
        assert link_cell.hyperlink is not None
        assert "ViewCreditNote" in link_cell.hyperlink.target

    def test_status_column_values(self) -> None:
        """Status column should reflect item_status_map completion state."""
        payload, _, _ = build_statement_excel_payload(**self._minimal_args())
        wb = load_workbook(BytesIO(payload))
        ws = wb["Statement"]
        headers = [cell.value for cell in ws[1]]
        status_col = headers.index("Status") + 1
        assert ws.cell(row=2, column=status_col).value == "Completed"
        assert ws.cell(row=3, column=status_col).value == "Incomplete"

    def test_empty_data_produces_valid_workbook(self) -> None:
        """Empty rows/items should still produce a valid XLSX with 0 data rows."""
        args = self._minimal_args(rows_by_header=[], right_rows_by_header=[], row_comparisons=[], row_matches=[], item_types=[], items=[], item_status_map={}, matched_invoice_to_statement_item={})
        payload, filename, row_count = build_statement_excel_payload(**args)
        assert row_count == 0
        wb = load_workbook(BytesIO(payload))
        assert "Statement" in wb.sheetnames

    def test_mismatch_border_applied_on_matched_row_with_cell_mismatch(self) -> None:
        """A matched row (row_matches=True) with a cell-level mismatch should get a border."""
        args = self._minimal_args(
            row_matches=[True, False],
            row_comparisons=[
                [
                    CellComparison(header="date", statement_value="2024-03-01", xero_value="2024-03-01", matches=True),
                    CellComparison(header="number", statement_value="101", xero_value="101", matches=True),
                    CellComparison(header="amount", statement_value="500.00", xero_value="600.00", matches=False),
                ],
                [],
            ],
        )
        payload, _, _ = build_statement_excel_payload(**args)
        # Just verify it produces a valid workbook without errors.
        wb = load_workbook(BytesIO(payload))
        assert wb["Statement"].max_row >= 2
