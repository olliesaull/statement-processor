"""Flask application for the statement processor service."""

import json
import os
import secrets
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from io import BytesIO
from typing import Any

import requests
from authlib.integrations.base_client.errors import OAuthError
from authlib.integrations.flask_client import OAuth
from flask import Flask, abort, jsonify, redirect, render_template, request, session, url_for
from flask_caching import Cache
from flask_session import Session
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

import cache_provider
from config import CLIENT_ID, CLIENT_SECRET, S3_BUCKET_NAME, STAGE
from core.contact_config_metadata import EXAMPLE_CONFIG, FIELD_DESCRIPTIONS
from core.get_contact_config import get_contact_config, set_contact_config
from core.item_classification import guess_statement_item_type
from core.models import StatementItem
from logger import logger
from sync import check_load_required, sync_data
from tenant_data_repository import TenantDataRepository, TenantStatus
from utils.auth import (
    active_tenant_required,
    block_when_loading,
    clear_session_is_set_cookie,
    has_cookie_consent,
    route_handler_logging,
    save_xero_oauth2_token,
    scope_str,
    set_session_is_set_cookie,
    xero_token_required,
)
from utils.dynamo import (
    add_statement_to_table,
    delete_statement_data,
    get_completed_statements,
    get_incomplete_statements,
    get_statement_item_status_map,
    get_statement_record,
    mark_statement_completed,
    persist_item_types_to_dynamo,
    set_all_statement_items_completed,
    set_statement_item_completed,
)
from utils.statement_view import build_right_rows, build_row_comparisons, get_date_format_from_config, get_number_separators_from_config, match_invoices_to_statement_items, prepare_display_mappings
from utils.storage import StatementJSONNotFoundError, fetch_json_statement, is_allowed_pdf, statement_json_s3_key, statement_pdf_s3_key, upload_statement_to_s3
from utils.workflows import start_textraction_state_machine
from xero_repository import get_contacts, get_credit_notes_by_contact, get_invoices_by_contact, get_payments_by_contact

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", os.urandom(16))

# Enable CSRF protection globally
csrf = CSRFProtect(app)

MAX_UPLOAD_MB = os.getenv("MAX_UPLOAD_MB", "10")
MAX_UPLOAD_BYTES = int(MAX_UPLOAD_MB) * 1024 * 1024
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES

os.makedirs(app.instance_path, exist_ok=True)
session_dir = os.path.join(app.instance_path, "flask_session")
os.makedirs(session_dir, exist_ok=True)
app.config.update(SESSION_TYPE="filesystem", SESSION_FILE_DIR=session_dir, SESSION_PERMANENT=False, SESSION_USE_SIGNER=True)
Session(app)

cache = Cache(app, config={"CACHE_TYPE": "SimpleCache", "CACHE_DEFAULT_TIMEOUT": 0})
cache_provider.set_cache(cache)


# Mirror selected config values in Flask app config for convenience
app.config["CLIENT_ID"] = CLIENT_ID
app.config["CLIENT_SECRET"] = CLIENT_SECRET

XERO_OIDC_METADATA_URL = os.getenv("XERO_OIDC_METADATA_URL", "https://identity.xero.com/.well-known/openid-configuration")

if STAGE == "prod":
    REDIRECT_URI = "https://cloudcathode.com/callback"
elif STAGE == "dev":
    REDIRECT_URI = "https://s7mznicnms.eu-west-1.awsapprunner.com/callback"
else:
    REDIRECT_URI = "http://localhost:8080/callback"

oauth = OAuth(app)
oauth.register(
    name="xero",
    client_id=CLIENT_ID,
    client_secret=CLIENT_SECRET,
    # Load endpoints + JWKS from OIDC metadata so Authlib can validate id_tokens.
    server_metadata_url=XERO_OIDC_METADATA_URL,
    # Reuse the existing scope string to keep requested permissions unchanged.
    client_kwargs={"scope": scope_str()},
)

_executor = ThreadPoolExecutor(max_workers=2)


DEFAULT_DECIMAL_SEPARATOR = "."
DEFAULT_THOUSANDS_SEPARATOR = ","
DECIMAL_SEPARATOR_OPTIONS = [(".", "Dot (.)"), (",", "Comma (,)")]
THOUSANDS_SEPARATOR_OPTIONS = [("", "None"), (",", "Comma (,)"), (".", "Dot (.)"), (" ", "Space ( )"), ("'", "Apostrophe (')")]
DECIMAL_SEPARATOR_VALUES = {opt[0] for opt in DECIMAL_SEPARATOR_OPTIONS}
THOUSANDS_SEPARATOR_VALUES = {opt[0] for opt in THOUSANDS_SEPARATOR_OPTIONS}
_ITEM_TYPE_LABELS: dict[str, str] = {"credit_note": "CRN", "invoice": "INV", "payment": "PMT"}


def _trigger_initial_sync_if_required(tenant_id: str | None) -> None:
    """Kick off an initial load if the tenant has no cached data yet."""
    if not tenant_id:
        return

    if check_load_required(tenant_id):
        oauth_token = session.get("xero_oauth2_token")
        if not oauth_token:
            logger.warning("Skipping background sync; missing OAuth token", tenant_id=tenant_id)
        else:
            _executor.submit(sync_data, tenant_id, TenantStatus.LOADING, oauth_token)


def _set_active_tenant(tenant_id: str | None) -> None:
    """Persist the selected tenant in the session."""
    tenants = session.get("xero_tenants", []) or []
    tenant_map = {t.get("tenantId"): t for t in tenants if t.get("tenantId")}
    if tenant_id and tenant_id in tenant_map:
        session["xero_tenant_id"] = tenant_id
        session["xero_tenant_name"] = tenant_map[tenant_id].get("tenantName")
        _trigger_initial_sync_if_required(tenant_id)
    else:
        session.pop("xero_tenant_id", None)
        session.pop("xero_tenant_name", None)


def _normalize_decimal_separator(value: str | None) -> str:
    """Coerce the decimal separator to a supported value."""
    if value in DECIMAL_SEPARATOR_VALUES:
        return value or DEFAULT_DECIMAL_SEPARATOR
    return DEFAULT_DECIMAL_SEPARATOR


def _normalize_thousands_separator(value: str | None) -> str:
    """Coerce the thousands separator to a supported value."""
    if value in THOUSANDS_SEPARATOR_VALUES:
        return value if value is not None else DEFAULT_THOUSANDS_SEPARATOR
    return DEFAULT_THOUSANDS_SEPARATOR


def _build_config_rows(cfg: dict[str, Any]) -> list[dict[str, Any]]:
    """Build table rows for canonical fields using existing config values."""
    # Flatten mapping sources: nested 'statement_items' + root-level keys.
    nested = cfg.get("statement_items") if isinstance(cfg, dict) else None
    nested = nested if isinstance(nested, dict) else {}
    flat: dict[str, Any] = {}
    flat.update(nested)
    allowed_keys = set(StatementItem.model_fields.keys())
    disallowed = {"raw", "statement_item_id"}
    if isinstance(cfg, dict):
        for k, v in cfg.items():
            if k in allowed_keys and k not in disallowed:
                flat[k] = v

    flat.pop("reference", None)
    flat.pop("item_type", None)

    # Canonical field order from the Pydantic model, prioritising config UI alignment.
    preferred_order = ["number", "total", "date", "due_date"]
    model_fields = [f for f in dict(StatementItem.model_fields) if f not in {"raw", "statement_item_id", "item_type"}]
    remaining_fields = [f for f in model_fields if f not in preferred_order]
    canonical_order = preferred_order + remaining_fields

    rows: list[dict[str, Any]] = []
    for f in canonical_order:
        if f in {"reference", "item_type"}:
            continue
        val = flat.get(f)
        if f == "total":
            values = [str(v) for v in val] if isinstance(val, list) else ([str(val)] if isinstance(val, str) else [""])
            rows.append({"field": f, "values": values or [""], "is_multi": True})
        else:
            values = [str(val)] if isinstance(val, str) else [""]
            rows.append({"field": f, "values": values, "is_multi": False})
    return rows


def _load_config_context(tenant_id: str | None, contact_lookup: dict[str, str], selected_contact_name: str) -> dict[str, Any]:
    """Load a contact config and return updates for the template context."""
    selected_contact_id = contact_lookup.get(selected_contact_name)
    logger.info("Config load submitted", tenant_id=tenant_id, contact_name=selected_contact_name, contact_id=selected_contact_id)

    updates: dict[str, Any] = {"selected_contact_name": selected_contact_name, "selected_contact_id": selected_contact_id}

    if not selected_contact_id:
        updates["error"] = "Please select a valid contact."
        logger.info("Config load failed", tenant_id=tenant_id, contact_name=selected_contact_name)
        return updates

    try:
        cfg = get_contact_config(tenant_id, selected_contact_id)
        updates["mapping_rows"] = _build_config_rows(cfg)
        updates["decimal_separator"] = _normalize_decimal_separator(str(cfg.get("decimal_separator", "")))
        updates["thousands_separator"] = _normalize_thousands_separator(str(cfg.get("thousands_separator", "")))
        updates["date_format"] = str(cfg.get("date_format") or "") if isinstance(cfg, dict) else ""
        logger.info("Config loaded", tenant_id=tenant_id, contact_id=selected_contact_id, keys=len(cfg) if isinstance(cfg, dict) else 0)
        return updates
    except KeyError:
        updates["mapping_rows"] = _build_config_rows({})
        updates["decimal_separator"] = DEFAULT_DECIMAL_SEPARATOR
        updates["thousands_separator"] = DEFAULT_THOUSANDS_SEPARATOR
        updates["date_format"] = ""
        updates["message"] = "No existing config found. You can create one below."
        logger.info("Config not found", tenant_id=tenant_id, contact_id=selected_contact_id)
        return updates
    except Exception as exc:
        updates["error"] = f"Failed to load config: {exc}"
        logger.info("Config load error", tenant_id=tenant_id, contact_id=selected_contact_id, error=exc)
        return updates


def _save_config_context(tenant_id: str | None, form: Any) -> dict[str, Any]:
    """Persist config edits and return updates for the template context."""
    selected_contact_id = form.get("contact_id")
    selected_contact_name = form.get("contact_name")
    logger.info("Config save submitted", tenant_id=tenant_id, contact_id=selected_contact_id, contact_name=selected_contact_name)

    updates: dict[str, Any] = {"selected_contact_id": selected_contact_id, "selected_contact_name": selected_contact_name}

    try:
        try:
            existing = get_contact_config(tenant_id, selected_contact_id)
        except KeyError:
            existing = {}
        posted_fields = [f for f in form.getlist("fields[]") if f]

        selected_decimal_separator = _normalize_decimal_separator(form.get("decimal_separator"))
        selected_thousands_separator = _normalize_thousands_separator(form.get("thousands_separator"))
        selected_date_format = (form.get("date_format") or "").strip()

        # Preserve any root keys not shown in the mapping editor.
        # Explicitly drop legacy 'statement_items' (we no longer store nested mappings).
        preserved = {k: v for k, v in existing.items() if k not in [*posted_fields, "statement_items"] and k not in {"reference", "item_type"}}

        new_map: dict[str, Any] = {}
        for f in posted_fields:
            if f == "total":
                total_vals = [v.strip() for v in form.getlist("map[total][]") if v.strip()]
                new_map["total"] = total_vals
            else:
                val = form.get(f"map[{f}]")
                new_map[f] = (val or "").strip()
        combined = {**preserved, **new_map, "date_format": selected_date_format, "decimal_separator": selected_decimal_separator, "thousands_separator": selected_thousands_separator}

        updates["decimal_separator"] = selected_decimal_separator
        updates["thousands_separator"] = selected_thousands_separator
        updates["date_format"] = selected_date_format

        # Validate required mappings before saving.
        number_value = (new_map.get("number") or "").strip()
        total_values = new_map.get("total") if isinstance(new_map.get("total"), list) else []
        if not number_value:
            updates["error"] = "The 'Number' field is mandatory. Please map the statement column that contains item numbers (e.g. invoice number)."
            updates["message"] = None
            updates["mapping_rows"] = _build_config_rows(combined)
        elif not total_values:
            updates["error"] = "The 'Total' field is mandatory. Please map at least one statement column with totals."
            updates["message"] = None
            updates["mapping_rows"] = _build_config_rows(combined)
        else:
            set_contact_config(tenant_id, selected_contact_id, combined)
            logger.info("Contact config saved", tenant_id=tenant_id, contact_id=selected_contact_id, contact_name=selected_contact_name, config=combined)
            updates["message"] = "Config updated successfully."
            updates["mapping_rows"] = _build_config_rows(combined)
    except Exception as exc:
        updates["error"] = f"Failed to save config: {exc}"
        logger.info("Config save failed", tenant_id=tenant_id, contact_id=selected_contact_id, error=exc)
    return updates


@app.route("/api/tenant-statuses", methods=["GET"])
@xero_token_required
def tenant_status():
    """Return tenant sync statuses and refresh cached status values."""
    tenant_records = session.get("xero_tenants", []) or []
    tenant_ids = [t.get("tenantId") for t in tenant_records if isinstance(t, dict)]
    try:
        tenant_statuses = TenantDataRepository.get_tenant_statuses(tenant_ids)
    except Exception as exc:
        logger.exception("Failed to load tenant sync status", tenant_ids=tenant_ids, error=exc)
        return jsonify({"error": "Unable to determine sync status"}), 500

    # Cache statuses so polling endpoints can respond quickly.
    for tenant_id, status in tenant_statuses.items():
        if tenant_id:
            cache_provider.set_tenant_status_cache(tenant_id, str(status))

    return jsonify(tenant_statuses), 200


@app.route("/api/tenants/<tenant_id>/sync", methods=["POST"])
@xero_token_required
def trigger_tenant_sync(tenant_id: str):
    """Trigger a background sync for the specified tenant."""
    tenant_id = (tenant_id or "").strip()
    if not tenant_id:
        return jsonify({"error": "TenantID is required"}), 400

    # Only allow syncs for tenants already connected in this session.
    tenant_records = session.get("xero_tenants", []) or []
    tenant_ids = {t.get("tenantId") for t in tenant_records if isinstance(t, dict)}
    if tenant_id not in tenant_ids:
        logger.info("Manual sync denied; tenant not authorized", tenant_id=tenant_id)
        return jsonify({"error": "Tenant not authorized"}), 403

    oauth_token = session.get("xero_oauth2_token")
    if not oauth_token:
        logger.warning("Manual sync denied; missing OAuth token", tenant_id=tenant_id)
        return jsonify({"error": "Missing OAuth token"}), 400

    try:
        # Fire-and-forget: sync runs in the background.
        _executor.submit(sync_data, tenant_id, TenantStatus.SYNCING, oauth_token)  # TODO: Perhaps worth checking if there is row in DDB/files in S3
        logger.info("Manual tenant sync triggered", tenant_id=tenant_id)
        return jsonify({"started": True}), 202
    except Exception as exc:
        logger.exception("Failed to trigger manual sync", tenant_id=tenant_id, error=exc)
        return jsonify({"error": "Failed to trigger sync"}), 500


@app.route("/")
@route_handler_logging
def index():
    """Render the landing page."""
    logger.info("Rendering index")
    return render_template("index.html")


@app.route("/tenant_management")
@route_handler_logging
@xero_token_required
def tenant_management():
    """Render tenant management, consuming one-time messages from session."""
    tenants = session.get("xero_tenants") or []
    active_tenant_id = session.get("xero_tenant_id")
    # Messages are popped so they only display once.
    message = session.pop("tenant_message", None)
    error = session.pop("tenant_error", None)

    active_tenant = next((t for t in tenants if t.get("tenantId") == active_tenant_id), None)
    logger.info(
        "Rendering tenant_management page", active_tenant_id=active_tenant_id, tenants=len(tenants), has_message=bool(message), has_error=bool(error), authenticated=bool(session.get("access_token"))
    )

    return render_template("tenant_management.html", tenants=tenants, active_tenant_id=active_tenant_id, active_tenant=active_tenant, message=message, error=error)


@app.route("/favicon.ico")
def ignore_favicon():
    """Return empty 204 for favicon requests."""
    return "", 204


def _get_active_contacts_for_upload() -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Return active contacts and a name -> ID lookup for the upload form."""
    contacts_raw = get_contacts()
    contacts_active = [c for c in contacts_raw if str(c.get("contact_status") or "").upper() == "ACTIVE"]
    contacts_list = sorted(contacts_active, key=lambda c: (c.get("name") or "").casefold())
    contact_lookup = {c["name"]: c["contact_id"] for c in contacts_list}
    return contacts_list, contact_lookup


def _validate_upload_payload(files: list, names: list[str]) -> bool:
    """Validate the number of uploaded files and selected contacts."""
    if not files:
        logger.info("Upload rejected; no statement files provided.")
        return False
    if len(files) != len(names):
        logger.info("Upload rejected; file count does not match contact selections.")
        return False
    return True


def _ensure_contact_config(tenant_id: str | None, contact_id: str, contact_name: str, filename: str, error_messages: list[str]) -> bool:
    """Ensure the contact has a config; on failure, log and append a user-facing error."""
    try:
        get_contact_config(tenant_id, contact_id)
    except KeyError:
        logger.warning("Upload blocked; contact config missing", tenant_id=tenant_id, contact_id=contact_id, contact_name=contact_name, statement_filename=filename)
        error_messages.append(f"Contact '{contact_name}' does not have a statement config yet. Please configure it before uploading.")
        return False
    except Exception as exc:
        logger.exception("Upload blocked; config lookup failed", tenant_id=tenant_id, contact_id=contact_id, contact_name=contact_name, statement_filename=filename, error=exc)
        error_messages.append(f"Could not load the config for '{contact_name}'. Please try again later.")
        return False
    return True


def _process_statement_upload(tenant_id: str | None, uploaded_file: FileStorage, contact_id: str, contact_name: str) -> str:
    """Upload the PDF, register the statement, and kick off textraction."""
    file_bytes = getattr(uploaded_file, "content_length", None)
    statement_id = str(uuid.uuid4())
    logger.info(
        "Preparing statement upload", tenant_id=tenant_id, contact_id=contact_id, contact_name=contact_name, statement_id=statement_id, statement_filename=uploaded_file.filename, bytes=file_bytes
    )

    entry = {"statement_id": statement_id, "statement_name": uploaded_file.filename, "contact_name": contact_name, "contact_id": contact_id}

    # Upload PDF to S3 first so downstream processing can read it.
    pdf_statement_key = statement_pdf_s3_key(tenant_id, statement_id)
    upload_statement_to_s3(fs_like=uploaded_file, key=pdf_statement_key)
    logger.info("Uploaded statement PDF", tenant_id=tenant_id, contact_id=contact_id, statement_id=statement_id, s3_key=pdf_statement_key)

    # Persist statement metadata to DynamoDB.
    add_statement_to_table(tenant_id, entry)
    logger.info("Statement submitted and metadata registered", tenant_id=tenant_id, contact_id=contact_id, statement_id=statement_id, table_entry=entry)

    # Kick off background textraction so it's ready by the time the user views it.
    json_statement_key = statement_json_s3_key(tenant_id, statement_id)
    started = start_textraction_state_machine(tenant_id=tenant_id, contact_id=contact_id, statement_id=statement_id, pdf_key=pdf_statement_key, json_key=json_statement_key)

    log_kwargs = {"tenant_id": tenant_id, "contact_id": contact_id, "statement_id": statement_id, "pdf_key": pdf_statement_key, "json_key": json_statement_key}

    if started:
        logger.info("Started textraction workflow", **log_kwargs)
    else:
        logger.error("Failed to start textraction workflow", **log_kwargs)

    return statement_id


@app.route("/upload-statements", methods=["GET", "POST"])
@active_tenant_required("Please select a tenant before uploading statements.")
@xero_token_required
@route_handler_logging
@block_when_loading
def upload_statements():
    """Upload one or more PDF statements and register them for processing."""
    tenant_id = session.get("xero_tenant_id")

    contacts_list, contact_lookup = _get_active_contacts_for_upload()
    success_count: int | None = None
    error_messages: list[str] = []
    logger.info("Rendering upload statements", tenant_id=tenant_id, available_contacts=len(contacts_list))

    uploads_ok = 0
    if request.method == "POST":
        files = [f for f in request.files.getlist("statements") if f and f.filename]
        names = request.form.getlist("contact_names")
        logger.info("Upload statements submitted", tenant_id=tenant_id, files=len(files), names=len(names))
        if _validate_upload_payload(files, names):
            for uploaded_file, contact in zip(files, names, strict=False):
                if not contact.strip():
                    logger.info("Missing contact", statement_filename=uploaded_file.filename)
                    continue
                if not is_allowed_pdf(uploaded_file.filename, uploaded_file.mimetype):
                    logger.info("Rejected non-PDF upload", statement_filename=uploaded_file.filename)
                    continue

                contact_name = contact.strip()
                contact_id: str | None = contact_lookup.get(contact_name)
                if not contact_id:
                    logger.warning("Upload blocked; contact not found", tenant_id=tenant_id, contact_name=contact_name, statement_filename=uploaded_file.filename)
                    error_messages.append(f"Contact '{contact_name}' was not recognised. Please select a contact from the list.")  # nosec B608 - user-facing message only, no SQL execution
                    continue

                if not _ensure_contact_config(tenant_id, contact_id, contact_name, uploaded_file.filename, error_messages):
                    continue

                _process_statement_upload(tenant_id=tenant_id, uploaded_file=uploaded_file, contact_id=contact_id, contact_name=contact_name)
                uploads_ok += 1

        if uploads_ok:
            success_count = uploads_ok
        logger.info("Upload statements processed", tenant_id=tenant_id, succeeded=uploads_ok, errors=list(error_messages))

    return render_template("upload_statements.html", contacts=contacts_list, success_count=success_count, error_messages=error_messages)


@app.route("/instructions")
@route_handler_logging
def instructions():
    """Render the user instructions page."""
    return render_template("instructions.html")


@app.route("/about")
@route_handler_logging
def about():
    """Render the about page."""
    return render_template("about.html")


@app.route("/cookies")
@route_handler_logging
def cookies():
    """Render the cookie policy and consent page."""
    return render_template("cookies.html")


@app.route("/statements")
@active_tenant_required("Please select a tenant to view statements.")
@xero_token_required
@route_handler_logging
@block_when_loading
def statements():
    """Render the statement list with filtering and sorting."""
    tenant_id = session.get("xero_tenant_id")

    # Read query params and normalize sort direction.
    view = request.args.get("view", "incomplete").lower()
    show_completed = view == "completed"
    statement_rows = get_completed_statements() if show_completed else get_incomplete_statements()
    sort_key = request.args.get("sort", "uploaded").lower()
    dir_param = (request.args.get("dir") or "").strip().lower()
    ALLOWED_DIR = {"asc", "desc"}
    default_dir_map = {"contact": "asc", "date_range": "desc", "uploaded": "desc"}
    if sort_key not in {"contact", "date_range", "uploaded"}:
        sort_key = "uploaded"
    current_dir = dir_param if dir_param in ALLOWED_DIR else default_dir_map.get(sort_key, "desc")
    reverse = current_dir == "desc"
    message = session.pop("statements_message", None)

    def _parse_iso_date(value: object) -> date | None:
        if not isinstance(value, str):
            return None
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return date.fromisoformat(stripped)
        except ValueError:
            return None

    def _parse_iso_datetime(value: object) -> datetime | None:
        if not isinstance(value, str):
            return None
        text = value.strip()
        if not text:
            return None
        try:
            # Support both "+00:00" and trailing "Z"
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None

    # Add derived fields for display and sorting.
    for row in statement_rows:
        earliest = _parse_iso_date(row.get("EarliestItemDate"))
        latest = _parse_iso_date(row.get("LatestItemDate"))
        row["_earliest_item_date"] = earliest
        row["_latest_item_date"] = latest
        row["_uploaded_at"] = _parse_iso_datetime(row.get("UploadedAt"))
        if earliest and latest:
            row["ItemDateRangeDisplay"] = earliest.isoformat() if earliest == latest else f"{earliest.isoformat()} - {latest.isoformat()}"
        elif latest:
            row["ItemDateRangeDisplay"] = latest.isoformat()
        elif earliest:
            row["ItemDateRangeDisplay"] = earliest.isoformat()
        else:
            row["ItemDateRangeDisplay"] = "—"

    if sort_key == "date_range":
        statement_rows.sort(key=lambda r: r.get("_latest_item_date") or date.min, reverse=reverse)
    elif sort_key == "uploaded":
        statement_rows.sort(key=lambda r: r.get("_uploaded_at") or datetime.min, reverse=reverse)
    else:
        # Contact: alphabetical or reverse, always keep missing/blank names last
        sort_key = "contact"
        nonempty = [r for r in statement_rows if isinstance(r.get("ContactName"), str) and r.get("ContactName").strip()]
        empty = [r for r in statement_rows if r not in nonempty]
        nonempty.sort(key=lambda r: str(r.get("ContactName")).strip().casefold(), reverse=reverse)
        statement_rows = nonempty + empty

    # Remove helper fields before rendering.
    for row in statement_rows:
        row.pop("_earliest_item_date", None)
        row.pop("_latest_item_date", None)
        row.pop("_uploaded_at", None)

    # Preserve filters when building sort URLs.
    base_args: dict[str, Any] = {}
    if show_completed:
        base_args["view"] = "completed"

    # For each sort key, clicking its button toggles the direction if already active,
    # otherwise applies the default direction for that key.
    def next_dir_for(key: str) -> str:
        if key == sort_key:
            return "asc" if current_dir == "desc" else "desc"
        return default_dir_map.get(key, "desc")

    sort_links = {
        "contact": url_for("statements", **dict(base_args, sort="contact", dir=next_dir_for("contact"))),
        "date_range": url_for("statements", **dict(base_args, sort="date_range", dir=next_dir_for("date_range"))),
        "uploaded": url_for("statements", **dict(base_args, sort="uploaded", dir=next_dir_for("uploaded"))),
    }

    logger.info("Rendering statements", tenant_id=tenant_id, view=view, sort=sort_key, direction=current_dir, statements=len(statement_rows))

    return render_template("statements.html", statements=statement_rows, show_completed=show_completed, message=message, current_sort=sort_key, current_dir=current_dir, sort_links=sort_links)


@app.route("/statement/<statement_id>/delete", methods=["POST"])
@active_tenant_required("Please select a tenant before deleting statements.")
@xero_token_required
@route_handler_logging
@block_when_loading
def delete_statement(statement_id: str):
    """Delete the statement and redirect back to the list view."""
    tenant_id = session.get("xero_tenant_id")

    try:
        delete_statement_data(tenant_id, statement_id)
        session["statements_message"] = "Statement deleted."
    except Exception as exc:
        logger.exception("Failed to delete statement", tenant_id=tenant_id, statement_id=statement_id, error=exc)
        session["tenant_error"] = "Unable to delete the statement. Please try again."

    return redirect(url_for("statements"))


def _parse_items_view(raw_value: str | None) -> str:
    """Normalize the statement item filter."""
    items_view = (raw_value or "incomplete").strip().lower()
    if items_view not in {"incomplete", "completed", "all"}:
        return "incomplete"
    return items_view


def _parse_show_payments(raw_value: str | None) -> bool:
    """Normalize the show payments flag."""
    value = (raw_value or "true").strip().lower()
    return value in {"true", "1", "yes", "on"}


def _handle_statement_post_actions(*, tenant_id: str, statement_id: str, form: Any, items_view: str, show_payments: bool) -> Any:
    """Handle POST actions for statement detail views, returning a redirect when applicable."""
    action = form.get("action")
    if action in {"mark_complete", "mark_incomplete"}:
        completed_flag = action == "mark_complete"
        try:
            mark_statement_completed(tenant_id, statement_id, completed_flag)
            try:
                set_all_statement_items_completed(tenant_id, statement_id, completed_flag)
            except Exception as exc:
                logger.exception("Failed to toggle all statement items", statement_id=statement_id, tenant_id=tenant_id, desired_state=completed_flag, error=exc)

            session["statements_message"] = "Statement marked as complete." if completed_flag else "Statement marked as incomplete."
            logger.info("Statement completion updated", tenant_id=tenant_id, statement_id=statement_id, completed=completed_flag)
        except Exception as exc:  # pragma: no cover - defensive logging
            logger.exception("Failed to toggle statement completion", statement_id=statement_id, tenant_id=tenant_id, desired_state=completed_flag, error=exc)
            abort(500)
        return redirect(url_for("statements"))

    if action in {"complete_item", "incomplete_item"}:
        statement_item_id = (form.get("statement_item_id") or "").strip()
        if statement_item_id:
            desired_state = action == "complete_item"
            try:
                set_statement_item_completed(tenant_id, statement_item_id, desired_state)
                logger.info("Statement item updated", tenant_id=tenant_id, statement_id=statement_id, statement_item_id=statement_item_id, completed=desired_state)
            except Exception as exc:
                logger.exception(
                    "Failed to toggle statement item completion", statement_id=statement_id, statement_item_id=statement_item_id, tenant_id=tenant_id, desired_state=desired_state, error=exc
                )
        return redirect(url_for("statement", statement_id=statement_id, items_view=items_view, show_payments="true" if show_payments else "false"))

    return None


def _build_match_by_item_id(matched_invoice_to_statement_item: dict[str, Any]) -> dict[str, dict[str, str]]:
    """Return a map of statement_item_id to matched document type/source."""
    match_by_item_id: dict[str, dict[str, str]] = {}
    for match in matched_invoice_to_statement_item.values():
        stmt_item = match.get("statement_item") if isinstance(match, dict) else None
        doc = match.get("invoice") if isinstance(match, dict) else None
        if not isinstance(stmt_item, dict) or not isinstance(doc, dict):
            continue
        statement_item_id = stmt_item.get("statement_item_id")
        if not statement_item_id:
            continue
        doc_type = str(doc.get("type") or "").upper()
        if doc.get("credit_note_id") or doc_type.endswith("CREDIT"):
            match_by_item_id[statement_item_id] = {"type": "credit_note", "source": "credit_note_match"}
        else:
            match_by_item_id[statement_item_id] = {"type": "invoice", "source": "invoice_match"}
    return match_by_item_id


def _build_payment_number_map(invoices: list[dict[str, Any]], payments: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Build a map of invoice number -> payment rows for payment inference."""
    invoice_number_by_id: dict[str, str] = {}
    for inv in invoices:
        inv_id = inv.get("invoice_id") if isinstance(inv, dict) else None
        inv_number = str(inv.get("number") or "").strip() if isinstance(inv, dict) else ""
        if inv_id and inv_number:
            invoice_number_by_id[str(inv_id)] = inv_number

    payment_number_map: dict[str, list[dict[str, Any]]] = {}
    for payment in payments:
        if not isinstance(payment, dict):
            continue
        invoice_id = payment.get("invoice_id")
        if not invoice_id:
            continue
        invoice_number = invoice_number_by_id.get(str(invoice_id))
        if not invoice_number:
            continue
        payment_number_map.setdefault(invoice_number, []).append(payment)
    return payment_number_map


def _classify_statement_items(
    *,
    items: list[Any],
    rows_by_header: list[dict[str, Any]],
    item_number_header: str | None,
    contact_config: dict[str, Any],
    matched_invoice_to_statement_item: dict[str, Any],
    matched_numbers: set[str],
    match_by_item_id: dict[str, dict[str, str]],
    payment_number_map: dict[str, list[dict[str, Any]]],
    statement_id: str,
) -> tuple[list[str], dict[str, str]]:
    """Classify statement items in-place and return item types + updates."""
    classification_updates: dict[str, str] = {}
    for idx, it in enumerate(items):
        if not isinstance(it, dict):
            continue
        statement_item_id = it.get("statement_item_id")
        raw = it.get("raw", {}) if isinstance(it.get("raw"), dict) else {}
        current_type = str(it.get("item_type") or "").strip().lower()
        row_number = ""
        if item_number_header and idx < len(rows_by_header):
            row_number = str(rows_by_header[idx].get(item_number_header) or "").strip()

        new_type: str | None
        source: str | None
        new_type = None
        source = None

        if statement_item_id and statement_item_id in match_by_item_id:
            entry = match_by_item_id[statement_item_id]
            new_type = entry["type"]
            source = entry["source"]
        elif row_number and row_number in matched_numbers:
            match = matched_invoice_to_statement_item.get(row_number)
            doc = match.get("invoice") if isinstance(match, dict) else None
            if isinstance(doc, dict):
                doc_type = str(doc.get("type") or "").upper()
                if doc.get("credit_note_id") or doc_type.endswith("CREDIT"):
                    new_type = "credit_note"
                    source = "credit_note_match"
                else:
                    new_type = "invoice"
                    source = "invoice_match"
        elif row_number and row_number not in matched_numbers and row_number in payment_number_map:
            new_type = "payment"
            source = "payment_match"

        if not new_type:
            new_type = guess_statement_item_type(raw, it.get("total"), contact_config)
            source = "heuristic"

        if new_type and new_type != current_type:
            it["item_type"] = new_type
            if statement_item_id:
                classification_updates[statement_item_id] = new_type
            logger.info("Statement item type updated", statement_id=statement_id, statement_item_id=statement_item_id, new_type=new_type, previous_type=current_type or "", source=source)

    item_types = [str((it.get("item_type") if isinstance(it, dict) else "") or "").strip().lower() for it in items]
    return item_types, classification_updates


def _persist_classification_updates(*, data: dict[str, Any], statement_id: str, tenant_id: str, json_statement_key: str, classification_updates: dict[str, str]) -> None:
    """Persist updated item types back to S3 and DynamoDB."""
    if not classification_updates:
        return

    try:
        json_payload = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        upload_statement_to_s3(BytesIO(json_payload), json_statement_key)
        logger.info("Persisted statement item types to S3", statement_id=statement_id, updated=len(classification_updates))
    except Exception as exc:
        logger.exception("Failed to persist statement JSON", statement_id=statement_id, error=str(exc))

    persist_item_types_to_dynamo(tenant_id, classification_updates)
    logger.info("Persisted statement item types to DynamoDB", statement_id=statement_id, updated=len(classification_updates))


def _build_row_matches(rows_by_header: list[dict[str, Any]], item_number_header: str | None, matched_invoice_to_statement_item: dict[str, Any], row_comparisons: list[list[Any]]) -> list[bool]:
    """Return the per-row match status for coloring and export."""
    if item_number_header:
        row_matches: list[bool] = []
        for r in rows_by_header:
            num = (r.get(item_number_header) or "").strip()
            row_matches.append(bool(num and matched_invoice_to_statement_item.get(num)))
        return row_matches

    # Fallback: if no number mapping, use strict all-cells match
    return [all(cell.matches for cell in row) for row in row_comparisons]


def _build_excel_headers(display_headers: list[str]) -> tuple[list[tuple[str, str]], list[str]]:
    """Build label pairs and the Excel header row."""
    header_labels: list[tuple[str, str]] = []
    statement_headers: list[str] = []
    xero_headers: list[str] = []

    for header in display_headers:
        label = (header or "").replace("_", " ").strip()
        label = label[0].upper() + label[1:] if label else header or ""
        header_labels.append((header, label))
        statement_headers.append(f"Statement {label}")
        xero_headers.append(f"Xero {label}")

    excel_headers = ["Type", *statement_headers, *xero_headers, "Xero Link", "Status"]
    return header_labels, excel_headers


def _add_excel_legend(workbook: Workbook, *, fill_success: PatternFill, fill_danger: PatternFill, fill_warning: PatternFill, mismatch_border: Border) -> None:
    """Add a legend sheet describing statement row styles."""
    legend = workbook.create_sheet(title="Legend")
    legend.column_dimensions["A"].width = 30
    legend.column_dimensions["B"].width = 18
    legend.append(["Legend", ""])
    legend["A1"].font = Font(bold=True)
    legend.append(["Match", ""])
    legend["B2"].fill = fill_success
    legend.append(["Mismatch", ""])
    legend["B3"].fill = fill_danger
    legend.append(["Flagged anomaly", ""])
    legend["B4"].fill = fill_warning
    legend.append(["Cell mismatch (matched rows)", ""])
    legend["B5"].border = mismatch_border


def _status_for_excel_row(item: Any, item_status_map: dict[str, bool]) -> tuple[str, bool]:
    """Return the status label and completion flag for an item."""
    statement_item_id = item.get("statement_item_id") if isinstance(item, dict) else None
    if statement_item_id:
        is_item_completed = item_status_map.get(statement_item_id, False)
        status_label = "Completed" if is_item_completed else "Incomplete"
        return status_label, is_item_completed
    return "", False


def _build_excel_row_values(header_labels: list[tuple[str, str]], left_row: dict[str, Any], right_row: dict[str, Any], item_types: list[str], idx: int) -> list[Any]:
    """Build Excel row values from statement/xero data."""
    item_type = item_types[idx] if idx < len(item_types) else ""
    row_values: list[Any] = [_format_item_type_label(item_type)]
    for src_header, _ in header_labels:
        left_value = left_row.get(src_header, "") if isinstance(left_row, dict) else ""
        row_values.append("" if left_value is None else left_value)

    for src_header, _ in header_labels:
        right_value = right_row.get(src_header, "") if isinstance(right_row, dict) else ""
        row_values.append("" if right_value is None else right_value)

    return row_values


def _is_anomalous_item(item: Any) -> bool:
    """Return True when the item has anomaly flags."""
    raw_flags = item.get("_flags") if isinstance(item, dict) else None
    flag_list = raw_flags if isinstance(raw_flags, list) else []
    return any(isinstance(flag, str) and flag.strip() in {"ml-outlier", "invalid-date"} for flag in flag_list)


def _row_fill_for_item(item: Any, row_match: bool, *, fill_warning: PatternFill, fill_success: PatternFill, fill_danger: PatternFill) -> PatternFill:
    """Return the fill color for the row."""
    if _is_anomalous_item(item):
        return fill_warning
    return fill_success if row_match else fill_danger


def _apply_row_fill(worksheet, *, current_row: int, total_columns: int, fill: PatternFill) -> None:
    """Apply row coloring to a worksheet row."""
    for col in range(1, total_columns + 1):
        cell = worksheet.cell(row=current_row, column=col)
        cell.fill = fill


def _apply_divider_borders(worksheet, *, current_row: int, statement_end_col: int, xero_start_col: int, divider_side: Side) -> None:
    """Apply divider borders between statement and Xero sections."""
    worksheet.cell(row=current_row, column=statement_end_col).border = Border(right=divider_side)
    worksheet.cell(row=current_row, column=xero_start_col).border = Border(left=divider_side)


def _apply_mismatch_borders(
    worksheet,
    *,
    header_labels: list[tuple[str, str]],
    comparisons: list[Any],
    current_row: int,
    statement_end_col: int,
    xero_start_col: int,
    mismatch_border: Border,
    mismatch_side: Side,
    divider_side: Side,
) -> None:
    """Apply per-cell mismatch borders for matched rows."""
    col_count = len(header_labels)
    for col_idx, comparison in enumerate(comparisons[:col_count]):
        if getattr(comparison, "matches", True):
            continue
        for target_col in (2 + col_idx, 2 + col_count + col_idx):
            cell = worksheet.cell(row=current_row, column=target_col)
            if target_col == statement_end_col:
                cell.border = Border(left=mismatch_side, right=divider_side, top=mismatch_side, bottom=mismatch_side)
            elif target_col == xero_start_col:
                cell.border = Border(left=divider_side, right=mismatch_side, top=mismatch_side, bottom=mismatch_side)
            else:
                cell.border = mismatch_border


def _parse_date_value(value: Any) -> date | None:
    """Parse a date value from a record field."""
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _format_date_segment(earliest_date: date | None, latest_date: date | None) -> str:
    """Return the filename date segment from the parsed dates."""
    if earliest_date and latest_date:
        if earliest_date == latest_date:
            return earliest_date.strftime("%Y-%m-%d")
        return f"{earliest_date.strftime('%Y-%m-%d')}_{latest_date.strftime('%Y-%m-%d')}"
    if latest_date or earliest_date:
        chosen = latest_date or earliest_date
        return chosen.strftime("%Y-%m-%d") if chosen else ""
    return ""


def _contact_segment(record: dict[str, Any], statement_id: str) -> str:
    """Return the filename contact segment."""
    contact_name = record.get("ContactName") if isinstance(record, dict) else ""
    return secure_filename(str(contact_name or "").strip()) or f"statement_{statement_id}"


def _append_excel_rows(
    worksheet,
    *,
    header_labels: list[tuple[str, str]],
    excel_headers: list[str],
    rows_by_header: list[dict[str, Any]],
    right_rows_by_header: list[dict[str, Any]],
    row_comparisons: list[list[Any]],
    row_matches: list[bool],
    item_types: list[str],
    items: list[Any],
    item_number_header: str | None,
    matched_invoice_to_statement_item: dict[str, Any],
    item_status_map: dict[str, bool],
    statement_col_count: int,
    statement_end_col: int,
    xero_start_col: int,
    fill_success: PatternFill,
    fill_danger: PatternFill,
    fill_warning: PatternFill,
    mismatch_border: Border,
    mismatch_side: Side,
    divider_side: Side,
) -> int:
    """Append rows to the Excel worksheet and return row count.

    This includes a hyperlink cell for the Xero Link column when available.
    """
    row_count = max(len(rows_by_header), len(right_rows_by_header))
    try:
        link_col = excel_headers.index("Xero Link") + 1
    except ValueError:
        link_col = None
    for idx in range(row_count):
        left_row = rows_by_header[idx] if idx < len(rows_by_header) else {}
        right_row = right_rows_by_header[idx] if idx < len(right_rows_by_header) else {}
        item = items[idx] if idx < len(items) else {}

        status_label, _is_item_completed = _status_for_excel_row(item, item_status_map)
        row_values = _build_excel_row_values(header_labels, left_row, right_row, item_types, idx)
        xero_invoice_id, xero_credit_note_id = _xero_ids_for_row(item_number_header, left_row, matched_invoice_to_statement_item)
        if xero_credit_note_id:
            xero_link = f"https://go.xero.com/AccountsPayable/ViewCreditNote.aspx?creditNoteID={xero_credit_note_id}"
        elif xero_invoice_id:
            xero_link = f"https://go.xero.com/AccountsPayable/View.aspx?InvoiceID={xero_invoice_id}"
        else:
            xero_link = ""
        # Providing status in the sheet lets users filter finished work out quickly.
        row_values.append("Link" if xero_link else "")
        row_values.append(status_label)
        worksheet.append(row_values)
        current_row = worksheet.max_row
        if xero_link and link_col:
            link_cell = worksheet.cell(row=current_row, column=link_col)
            link_cell.hyperlink = xero_link

        row_match = row_matches[idx] if idx < len(row_matches) else False
        fill = _row_fill_for_item(item, row_match, fill_warning=fill_warning, fill_success=fill_success, fill_danger=fill_danger)
        _apply_row_fill(worksheet, current_row=current_row, total_columns=len(excel_headers), fill=fill)

        if statement_col_count:
            _apply_divider_borders(worksheet, current_row=current_row, statement_end_col=statement_end_col, xero_start_col=xero_start_col, divider_side=divider_side)

        if row_match and idx < len(row_comparisons):
            comparisons = row_comparisons[idx] or []
            _apply_mismatch_borders(
                worksheet,
                header_labels=header_labels,
                comparisons=comparisons,
                current_row=current_row,
                statement_end_col=statement_end_col,
                xero_start_col=xero_start_col,
                mismatch_border=mismatch_border,
                mismatch_side=mismatch_side,
                divider_side=divider_side,
            )
    return row_count


def _build_statement_excel_response(
    *,
    display_headers: list[str],
    rows_by_header: list[dict[str, Any]],
    right_rows_by_header: list[dict[str, Any]],
    row_comparisons: list[list[Any]],
    row_matches: list[bool],
    item_types: list[str],
    items: list[Any],
    item_number_header: str | None,
    matched_invoice_to_statement_item: dict[str, Any],
    item_status_map: dict[str, bool],
    record: dict[str, Any],
    statement_id: str,
    tenant_id: str,
) -> Any:
    """Build an XLSX export response for the current statement view.
    This applies basic header formatting plus a frozen header row with dropdown filters.

    Args:
        display_headers: Statement display headers.
        rows_by_header: Statement rows keyed by header.
        right_rows_by_header: Xero rows keyed by header.
        row_comparisons: Per-cell comparison results.
        row_matches: Per-row match flags.
        item_types: Item type labels per row.
        items: Statement items payload.
        item_number_header: Header used to map Xero links.
        matched_invoice_to_statement_item: Matched Xero invoice map.
        item_status_map: Statement item completion flags.
        record: Statement metadata record.
        statement_id: Statement identifier.
        tenant_id: Active tenant identifier.

    Returns:
        Flask response containing the XLSX export.
    """
    header_labels, excel_headers = _build_excel_headers(display_headers)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Statement"
    worksheet.append(excel_headers)

    fill_success = PatternFill(fill_type="solid", fgColor="C6EFCE")
    fill_danger = PatternFill(fill_type="solid", fgColor="CD5C5C")
    fill_warning = PatternFill(fill_type="solid", fgColor="FFEB9C")
    mismatch_side = Side(style="thin", color="D8A0A0")
    mismatch_border = Border(left=mismatch_side, right=mismatch_side, top=mismatch_side, bottom=mismatch_side)
    divider_side = Side(style="medium", color="808080")
    statement_col_count = len(header_labels)
    statement_end_col = 1 + statement_col_count
    xero_start_col = statement_end_col + 1

    _add_excel_legend(workbook, fill_success=fill_success, fill_danger=fill_danger, fill_warning=fill_warning, mismatch_border=mismatch_border)

    if statement_col_count:
        worksheet.cell(row=1, column=statement_end_col).border = Border(right=divider_side)
        worksheet.cell(row=1, column=xero_start_col).border = Border(left=divider_side)

    row_count = _append_excel_rows(
        worksheet,
        header_labels=header_labels,
        excel_headers=excel_headers,
        rows_by_header=rows_by_header,
        right_rows_by_header=right_rows_by_header,
        row_comparisons=row_comparisons,
        row_matches=row_matches,
        item_types=item_types,
        items=items,
        item_number_header=item_number_header,
        matched_invoice_to_statement_item=matched_invoice_to_statement_item,
        item_status_map=item_status_map,
        statement_col_count=statement_col_count,
        statement_end_col=statement_end_col,
        xero_start_col=xero_start_col,
        fill_success=fill_success,
        fill_danger=fill_danger,
        fill_warning=fill_warning,
        mismatch_border=mismatch_border,
        mismatch_side=mismatch_side,
        divider_side=divider_side,
    )

    header_font = Font(bold=True)
    for col_idx in range(1, len(excel_headers) + 1):
        worksheet.cell(row=1, column=col_idx).font = header_font

    worksheet.freeze_panes = "A2"
    last_row = max(row_count + 1, 1)
    last_column = get_column_letter(len(excel_headers))
    worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    width_overrides = {"Type": 8, "Status": 12, "Xero Link": 12}
    for col_idx, header in enumerate(excel_headers, start=1):
        width = width_overrides.get(header)
        if width is None:
            width = min(max(len(header) + 2, 14), 30)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    excel_payload = output.getvalue()
    output.close()

    response = app.response_class(excel_payload, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    earliest_date = _parse_date_value(record.get("EarliestItemDate"))
    latest_date = _parse_date_value(record.get("LatestItemDate"))
    date_segment = _format_date_segment(earliest_date, latest_date)

    contact_segment = _contact_segment(record, statement_id)

    parts = [contact_segment]
    if date_segment:
        parts.append(date_segment)
    download_name = "_".join(parts) + "_export.xlsx"

    response.headers["Content-Disposition"] = f'attachment; filename="{download_name}"'
    logger.info("Statement Excel generated", tenant_id=tenant_id, statement_id=statement_id, rows=row_count, excel_filename=download_name)
    return response


def _item_status(item: Any, item_status_map: dict[str, bool]) -> tuple[str | None, bool]:
    """Return the statement item ID and completion status."""
    statement_item_id = item.get("statement_item_id") if isinstance(item, dict) else None
    if statement_item_id:
        return statement_item_id, item_status_map.get(statement_item_id, False)
    return None, False


def _item_flags(item: Any) -> list[str]:
    """Return normalized, unique flags for a statement item."""
    if not isinstance(item, dict):
        return []
    raw_flags = item.get("_flags") or []
    if not isinstance(raw_flags, list):
        return []
    seen_flags: set[str] = set()
    flags: list[str] = []
    for flag in raw_flags:
        if not isinstance(flag, str):
            continue
        normalized = flag.strip()
        if not normalized or normalized in seen_flags:
            continue
        seen_flags.add(normalized)
        flags.append(normalized)
    return flags


def _format_item_type_label(item_type: str | None) -> str:
    """Format a statement item type for display.

    Args:
        item_type: Raw statement item type value.

    Returns:
        Display label for the item type.
    """
    normalized = str(item_type or "").strip().lower()
    if not normalized:
        return ""
    if normalized in _ITEM_TYPE_LABELS:
        return _ITEM_TYPE_LABELS[normalized]
    return normalized.replace("_", " ").upper()


def _xero_ids_for_row(item_number_header: str | None, left_row: dict[str, Any], matched_invoice_to_statement_item: dict[str, Any]) -> tuple[str | None, str | None]:
    """Return matched Xero invoice/credit note IDs for a row."""
    if not item_number_header:
        return None, None
    row_number = str(left_row.get(item_number_header) or "").strip()
    if not row_number:
        return None, None
    match = matched_invoice_to_statement_item.get(row_number)
    if not isinstance(match, dict):
        return None, None
    inv = match.get("invoice")
    if not isinstance(inv, dict):
        return None, None
    credit_note_id = inv.get("credit_note_id")
    xero_credit_note_id = credit_note_id.strip() if isinstance(credit_note_id, str) and credit_note_id.strip() else None
    invoice_id = inv.get("invoice_id")
    xero_invoice_id = invoice_id.strip() if isinstance(invoice_id, str) and invoice_id.strip() else None
    return xero_invoice_id, xero_credit_note_id


def _build_statement_rows(
    *,
    rows_by_header: list[dict[str, Any]],
    right_rows_by_header: list[dict[str, Any]],
    display_headers: list[str],
    row_comparisons: list[list[Any]],
    row_matches: list[bool],
    items: list[Any],
    item_types: list[str],
    item_status_map: dict[str, bool],
    item_number_header: str | None,
    matched_invoice_to_statement_item: dict[str, Any],
) -> list[dict[str, Any]]:
    """Build the rows displayed in the statement detail UI.

    Args:
        rows_by_header: Statement rows keyed by header names.
        right_rows_by_header: Xero rows keyed by header names.
        display_headers: Ordered list of display headers.
        row_comparisons: Per-cell comparison results.
        row_matches: Per-row match flags.
        items: Statement item payloads.
        item_types: Item type labels per row.
        item_status_map: Statement item completion flags.
        item_number_header: Header used to map Xero links.
        matched_invoice_to_statement_item: Matched Xero invoice map.

    Returns:
        List of row dicts for the statement detail table.
    """
    statement_rows: list[dict[str, Any]] = []
    for idx, left_row in enumerate(rows_by_header):
        item = items[idx] if idx < len(items) else {}
        statement_item_id, is_item_completed = _item_status(item, item_status_map)

        right_row_dict = right_rows_by_header[idx] if idx < len(right_rows_by_header) else {}
        flags = _item_flags(item)

        # Build Xero links by extracting IDs from matched data
        xero_invoice_id, xero_credit_note_id = _xero_ids_for_row(item_number_header, left_row, matched_invoice_to_statement_item)

        item_type = (item.get("item_type") if isinstance(item, dict) else None) or (item_types[idx] if idx < len(item_types) else "invoice")
        statement_rows.append(
            {
                "statement_item_id": statement_item_id,
                "left_values": [left_row.get(h, "") for h in display_headers],
                "right_values": [right_row_dict.get(h, "") for h in display_headers],
                "cell_comparisons": row_comparisons[idx] if idx < len(row_comparisons) else [],
                "matches": row_matches[idx] if idx < len(row_matches) else False,
                "is_completed": is_item_completed,
                "flags": flags,
                "item_type": item_type,
                "item_type_label": _format_item_type_label(item_type),
                "xero_invoice_id": xero_invoice_id,
                "xero_credit_note_id": xero_credit_note_id,
            }
        )

    return statement_rows


@app.route("/statement/<statement_id>", methods=["GET", "POST"])
@active_tenant_required("Please select a tenant to view statements.")
@xero_token_required
@route_handler_logging
@block_when_loading
def statement(statement_id: str):
    """Render the statement detail view, handling actions and exports."""
    tenant_id = session.get("xero_tenant_id")

    record = get_statement_record(tenant_id, statement_id)
    if not record:
        logger.info("Statement record not found", tenant_id=tenant_id, statement_id=statement_id)
        abort(404)

    items_view = _parse_items_view(request.values.get("items_view"))
    show_payments = _parse_show_payments(request.values.get("show_payments"))
    logger.info("Statement detail requested", tenant_id=tenant_id, statement_id=statement_id, items_view=items_view, show_payments=show_payments, method=request.method)

    raw_contact_name = record.get("ContactName")
    contact_name = str(raw_contact_name).strip() if raw_contact_name is not None else ""
    page_heading = contact_name or f"Statement {statement_id}"

    if request.method == "POST":
        response = _handle_statement_post_actions(tenant_id=tenant_id, statement_id=statement_id, form=request.form, items_view=items_view, show_payments=show_payments)
        if response is not None:
            return response

    json_statement_key = statement_json_s3_key(tenant_id, statement_id)

    contact_id = record.get("ContactID")
    is_completed = str(record.get("Completed", "")).lower() == "true"
    base_context: dict[str, Any] = {
        "statement_id": statement_id,
        "contact_name": contact_name,
        "page_heading": page_heading,
        "items_view": items_view,
        "show_payments": show_payments,
        "is_completed": is_completed,
    }
    try:
        data, _ = fetch_json_statement(tenant_id=tenant_id, contact_id=contact_id, bucket=S3_BUCKET_NAME, json_key=json_statement_key)
    except StatementJSONNotFoundError:
        logger.info("Statement JSON pending", tenant_id=tenant_id, statement_id=statement_id, json_key=json_statement_key)
        return render_template(
            "statement.html", is_processing=True, incomplete_count=0, completed_count=0, all_statement_rows=[], statement_rows=[], raw_statement_headers=[], has_payment_rows=False, **base_context
        )

    # 1) Parse display configuration and left-side rows
    items = data.get("statement_items", []) or []
    contact_config = get_contact_config(tenant_id, contact_id)
    decimal_sep, thousands_sep = get_number_separators_from_config(contact_config)
    display_headers, rows_by_header, header_to_field, item_number_header = prepare_display_mappings(items, contact_config)

    # 2) Fetch Xero documents and classify each statement item
    invoices = get_invoices_by_contact(contact_id) or []
    credit_notes = get_credit_notes_by_contact(contact_id) or []
    payments = get_payments_by_contact(contact_id) or []
    logger.info("Fetched Xero documents", statement_id=statement_id, contact_id=contact_id, invoices=len(invoices), credit_notes=len(credit_notes), payments=len(payments))

    docs_for_matching = invoices + credit_notes
    matched_invoice_to_statement_item = match_invoices_to_statement_items(items=items, rows_by_header=rows_by_header, item_number_header=item_number_header, invoices=docs_for_matching)

    matched_numbers: set[str] = {key for key in matched_invoice_to_statement_item if isinstance(key, str)}
    match_by_item_id = _build_match_by_item_id(matched_invoice_to_statement_item)
    payment_number_map = _build_payment_number_map(invoices, payments)

    item_types, classification_updates = _classify_statement_items(
        items=items,
        rows_by_header=rows_by_header,
        item_number_header=item_number_header,
        contact_config=contact_config,
        matched_invoice_to_statement_item=matched_invoice_to_statement_item,
        matched_numbers=matched_numbers,
        match_by_item_id=match_by_item_id,
        payment_number_map=payment_number_map,
        statement_id=statement_id,
    )

    _persist_classification_updates(data=data, statement_id=statement_id, tenant_id=tenant_id, json_statement_key=json_statement_key, classification_updates=classification_updates)

    # 3) Build right-hand rows from the matched invoices
    date_fmt = get_date_format_from_config(contact_config)

    right_rows_by_header = build_right_rows(
        rows_by_header=rows_by_header,
        display_headers=display_headers,
        header_to_field=header_to_field,
        matched_map=matched_invoice_to_statement_item,
        item_number_header=item_number_header,
        date_format=date_fmt,
        decimal_separator=decimal_sep,
        thousands_separator=thousands_sep,
    )

    # 4) Compare LEFT (statement) vs RIGHT (Xero) for per-cell indicators
    row_comparisons = build_row_comparisons(left_rows=rows_by_header, right_rows=right_rows_by_header, display_headers=display_headers, header_to_field=header_to_field)
    # Row highlight: if this row is linked to a Xero document (exact or substring),
    # consider the row a "match" for coloring purposes even if some cells differ.
    row_matches = _build_row_matches(rows_by_header, item_number_header, matched_invoice_to_statement_item, row_comparisons)

    item_status_map = get_statement_item_status_map(tenant_id, statement_id)

    if request.args.get("download") == "xlsx":
        return _build_statement_excel_response(
            display_headers=display_headers,
            rows_by_header=rows_by_header,
            right_rows_by_header=right_rows_by_header,
            row_comparisons=row_comparisons,
            row_matches=row_matches,
            item_types=item_types,
            items=items,
            item_number_header=item_number_header,
            matched_invoice_to_statement_item=matched_invoice_to_statement_item,
            item_status_map=item_status_map,
            record=record,
            statement_id=statement_id,
            tenant_id=tenant_id,
        )

    statement_rows = _build_statement_rows(
        rows_by_header=rows_by_header,
        right_rows_by_header=right_rows_by_header,
        display_headers=display_headers,
        row_comparisons=row_comparisons,
        row_matches=row_matches,
        items=items,
        item_types=item_types,
        item_status_map=item_status_map,
        item_number_header=item_number_header,
        matched_invoice_to_statement_item=matched_invoice_to_statement_item,
    )

    completed_count = sum(1 for row in statement_rows if row["is_completed"])
    incomplete_count = len(statement_rows) - completed_count
    has_payment_rows = any(row.get("item_type") == "payment" for row in statement_rows)

    if items_view == "completed":
        visible_rows = [row for row in statement_rows if row["is_completed"]]
    elif items_view == "incomplete":
        visible_rows = [row for row in statement_rows if not row["is_completed"]]
    else:
        visible_rows = statement_rows

    if not show_payments:
        visible_rows = [row for row in visible_rows if row.get("item_type") != "payment"]

    logger.info(
        "Statement detail rendered",
        tenant_id=tenant_id,
        statement_id=statement_id,
        visible=len(visible_rows),
        total=len(statement_rows),
        completed=completed_count,
        incomplete=incomplete_count,
        items_view=items_view,
        show_payments=show_payments,
    )

    context: dict[str, Any] = {
        **base_context,
        "is_processing": False,
        "raw_statement_headers": display_headers,
        "statement_rows": visible_rows,
        "all_statement_rows": statement_rows,
        "row_comparisons": row_comparisons,
        "completed_count": completed_count,
        "incomplete_count": incomplete_count,
        "has_payment_rows": has_payment_rows,
    }
    return render_template("statement.html", **context)


@app.route("/tenants/select", methods=["POST"])
@xero_token_required
@route_handler_logging
def select_tenant():
    """Persist the selected tenant in session and return to management view."""
    tenant_id = (request.form.get("tenant_id") or "").strip()
    tenants = session.get("xero_tenants") or []
    logger.info("Tenant selection submitted", tenant_id=tenant_id, available=len(tenants))

    if tenant_id and any(t.get("tenantId") == tenant_id for t in tenants):
        # Update the active tenant and display a success message.
        _set_active_tenant(tenant_id)
        tenant_name = session.get("xero_tenant_name") or tenant_id
        session["tenant_message"] = f"Switched to tenant: {tenant_name}."
        logger.info("Tenant switched", tenant_id=tenant_id, tenant_name=tenant_name)
    else:
        session["tenant_error"] = "Unable to select tenant. Please try again."
        logger.info("Tenant selection failed", tenant_id=tenant_id)

    return redirect(url_for("tenant_management"))


@app.route("/tenants/disconnect", methods=["POST"])
@xero_token_required
@route_handler_logging
def disconnect_tenant():
    """Disconnect the tenant from Xero and update the local session state."""
    tenant_id = (request.form.get("tenant_id") or "").strip()
    tenants = session.get("xero_tenants") or []
    tenant = next((t for t in tenants if t.get("tenantId") == tenant_id), None)
    management_url = url_for("tenant_management")

    if not tenant:
        session["tenant_error"] = "Tenant not found in session."
        return redirect(management_url)

    connection_id = tenant.get("connectionId")
    access_token = session.get("access_token")
    logger.info("Tenant disconnect submitted", tenant_id=tenant_id, has_connection=bool(connection_id))

    if connection_id and access_token:
        try:
            resp = requests.delete(f"https://api.xero.com/connections/{connection_id}", headers={"Authorization": f"Bearer {access_token}"}, timeout=20)
            if resp.status_code not in (200, 204):
                logger.error("Failed to disconnect tenant", tenant_id=tenant_id, status_code=resp.status_code, body=resp.text)
                session["tenant_error"] = "Unable to disconnect tenant from Xero."
                return redirect(management_url)
        except Exception as exc:
            logger.exception("Exception disconnecting tenant", tenant_id=tenant_id, error=exc)
            session["tenant_error"] = "An error occurred while disconnecting the tenant."
            return redirect(management_url)

    # Remove tenant locally regardless (in case it was already disconnected).
    updated = [t for t in tenants if t.get("tenantId") != tenant_id]
    session["xero_tenants"] = updated

    if session.get("xero_tenant_id") == tenant_id:
        next_tenant_id = updated[0]["tenantId"] if updated else None
        _set_active_tenant(next_tenant_id)

    session["tenant_message"] = "Tenant disconnected."
    logger.info("Tenant disconnected", tenant_id=tenant_id, remaining=len(updated))
    if not updated:
        return redirect(url_for("index"))
    return redirect(management_url)


@app.route("/configs", methods=["GET", "POST"])
@active_tenant_required("Please select a tenant before configuring mappings.")
@xero_token_required
@route_handler_logging
@block_when_loading
def configs():
    """Render and update the contact mapping configuration UI."""
    tenant_id = session.get("xero_tenant_id")

    contacts_raw = get_contacts()
    contacts_active = [c for c in contacts_raw if str(c.get("contact_status") or "").upper() == "ACTIVE"]
    contacts_list = sorted(contacts_active, key=lambda c: (c.get("name") or "").casefold())
    contact_lookup = {c["name"]: c["contact_id"] for c in contacts_list}
    logger.info("Rendering configs", tenant_id=tenant_id, contacts=len(contacts_list))

    context: dict[str, Any] = {
        "contacts": contacts_list,
        "selected_contact_name": None,
        "selected_contact_id": None,
        "mapping_rows": [],  # {field, values:list[str], is_multi:bool}
        "message": None,
        "error": None,
        "field_descriptions": FIELD_DESCRIPTIONS,
        "date_format": "",
        "decimal_separator": DEFAULT_DECIMAL_SEPARATOR,
        "thousands_separator": DEFAULT_THOUSANDS_SEPARATOR,
        "decimal_separator_options": DECIMAL_SEPARATOR_OPTIONS,
        "thousands_separator_options": THOUSANDS_SEPARATOR_OPTIONS,
    }

    if request.method == "POST":
        action = request.form.get("action")
        if action == "load":
            # Load existing config for the chosen contact name.
            selected_contact_name = (request.form.get("contact_name") or "").strip()
            context.update(_load_config_context(tenant_id, contact_lookup, selected_contact_name))
        elif action == "save_map":
            # Save edited mapping.
            context.update(_save_config_context(tenant_id, request.form))

    context.update(
        {
            "example_rows": _build_config_rows(EXAMPLE_CONFIG),
            "example_date_format": str(EXAMPLE_CONFIG.get("date_format") or ""),
            "example_decimal_separator": EXAMPLE_CONFIG.get("decimal_separator", DEFAULT_DECIMAL_SEPARATOR),
            "example_thousands_separator": EXAMPLE_CONFIG.get("thousands_separator", DEFAULT_THOUSANDS_SEPARATOR),
            "decimal_separator_labels": dict(DECIMAL_SEPARATOR_OPTIONS),
            "thousands_separator_labels": dict(THOUSANDS_SEPARATOR_OPTIONS),
        }
    )

    return render_template("configs.html", **context)


@app.route("/login")
@route_handler_logging
def login():
    """Start the Xero OAuth flow and redirect to the authorize URL."""
    logger.info("Login initiated")
    if not has_cookie_consent():
        logger.info("Login blocked; cookie consent missing")
        return redirect(url_for("cookies"))

    if not CLIENT_ID or not CLIENT_SECRET:
        return "Missing XERO_CLIENT_ID or XERO_CLIENT_SECRET env vars", 500

    # OIDC nonce ties the auth response to this browser session.
    nonce = secrets.token_urlsafe(24)
    session["oauth_nonce"] = nonce

    logger.info("Redirecting to Xero authorization", scope_count=len(scope_str().split()))
    # Authlib stores state/nonce in session and builds the authorize URL.
    return oauth.xero.authorize_redirect(redirect_uri=REDIRECT_URI, nonce=nonce)


@app.route("/callback")
@route_handler_logging
def callback():  # pylint: disable=too-many-return-statements
    """Handle the OAuth callback, validate tokens, and load tenant context."""
    if not has_cookie_consent():
        logger.info("OAuth callback blocked; cookie consent missing")
        return redirect(url_for("cookies"))

    # Handle user-denied or error cases
    error = request.args.get("error")
    if error is not None:
        error_description = request.args.get("error_description") or error
        logger.error("OAuth error", error_code=400, error_description=error_description, error=error)
        return f"OAuth error: {error_description}", 400

    try:
        tokens = oauth.xero.authorize_access_token()
    except OAuthError as exc:
        error_description = exc.description or exc.error
        logger.error("OAuth error", error_code=400, error_description=error_description, error=exc.error)
        return f"OAuth error: {error_description}", 400

    if not isinstance(tokens, dict):
        logger.error("Invalid token response from Xero", error_code=400)
        return "Invalid token response from Xero", 400

    # id_token is required for OIDC claim + nonce validation.
    if not tokens.get("id_token"):
        logger.error("Missing id_token in OAuth response", error_code=400)
        return "Missing id_token in OAuth response", 400

    nonce = session.pop("oauth_nonce", None)
    # Require the original nonce so we can validate the id_token against it.
    if not nonce:
        logger.error("Missing OAuth nonce in session", error_code=400)
        return "Missing OAuth nonce in session", 400

    try:
        # Validates signature + standard claims and checks nonce matches session.
        oauth.xero.parse_id_token(tokens, nonce=nonce)
    except Exception as exc:
        logger.exception("Failed to validate id_token", error=str(exc))
        return "Invalid id_token", 400

    save_xero_oauth2_token(tokens)
    access_token = tokens.get("access_token")
    session["access_token"] = access_token

    conn_res = requests.get("https://api.xero.com/connections", headers={"Authorization": f"Bearer {access_token}"}, timeout=20)

    conn_res.raise_for_status()
    connections = conn_res.json()
    if not connections:
        logger.error("No Xero connections found for this user.", error_code=400)
        return "No Xero connections found for this user.", 400

    tenants = [
        {"tenantId": conn.get("tenantId"), "tenantName": conn.get("tenantName"), "tenantType": conn.get("tenantType"), "connectionId": conn.get("id")} for conn in connections if conn.get("tenantId")
    ]

    current = session.get("xero_tenant_id")
    tenant_ids = [t["tenantId"] for t in tenants]

    # Store the latest tenant list before updating the active tenant.
    session["xero_tenants"] = tenants

    for tid in tenant_ids:
        _trigger_initial_sync_if_required(tid)

    if current in tenant_ids:
        _set_active_tenant(current)
    elif tenant_ids:
        first_tenant = tenant_ids[0]
        _set_active_tenant(first_tenant)
    else:
        _set_active_tenant(None)

    logger.info("OAuth callback processed", tenants=len(tenants))
    response = redirect(url_for("tenant_management"))
    return set_session_is_set_cookie(response)


@app.route("/logout")
@route_handler_logging
def logout():
    """Clear the session and return to the landing page."""
    logger.info("Logout requested", had_tenant=bool(session.get("xero_tenant_id")))
    session.clear()
    response = redirect(url_for("index"))
    return clear_session_is_set_cookie(response)


@app.route("/.well-known/<path:path>")
def chrome_devtools_ping(path):
    """Respond to Chrome DevTools well-known probes without logging 404s."""
    return "", 204  # No content, indicates "OK but nothing here"


if __name__ == "__main__":
    app.run(port=8080)
