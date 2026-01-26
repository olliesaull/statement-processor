"""Excel parsing helpers for Playwright tests."""

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from playwright_tests.helpers.runs import StatementFlowRun
from playwright_tests.helpers.tables import TableData


def _collapse_whitespace(value: str) -> str:
    """Normalize whitespace for string comparison.

    Args:
        value: Raw string value.

    Returns:
        Normalized string with collapsed whitespace.
    """
    return " ".join(value.split())


def _cell_to_text(cell: Cell) -> str:
    """Render a cell value as display text.

    Args:
        cell: OpenPyXL cell instance.

    Returns:
        Display string for the cell.
    """
    if cell.value is None:
        return ""
    return str(cell.value)


def read_excel_table(path: Path) -> TableData:
    """Read a worksheet into a TableData structure.

    Args:
        path: Path to the Excel file.

    Returns:
        TableData containing headers and rows.
    """
    workbook = load_workbook(path, data_only=True)

    rows: list[list[str]] = []
    for row in workbook.active.iter_rows():
        values = [_collapse_whitespace(_cell_to_text(cell)) for cell in row]
        if any(values):
            rows.append(values)

    if not rows:
        return TableData(headers=[], rows=[])

    headers = rows[0]
    data_rows = rows[1:]
    return TableData(headers=headers, rows=data_rows)


def require_expected_excel(test_run: StatementFlowRun) -> Path:
    """Return the expected Excel path or skip the test.

    Args:
        test_run: Current statement test run.

    Returns:
        Path to the expected Excel baseline.
    """
    expected_path = test_run.expected_excel_path()
    if expected_path is None:
        pytest.skip("expected_excel_filename is not set for this run.")
    if not expected_path.exists():
        pytest.skip(f"Expected Excel baseline not found: {expected_path}.")
    return expected_path
